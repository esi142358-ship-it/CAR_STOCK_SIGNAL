import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import STOCKS, STOCKS_NEXT50, STOCKS_SMALLCAP, ETF, SECTOR_MAP

OUTPUT_FILE = "output/Trading_System.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY_HEX    = "1B2A4A"
TEAL_HEX    = "0D7377"
PURPLE_HEX  = "4A0E8F"
LGREEN_HEX  = "C8F5D8"
LRED_HEX    = "FDDEDE"
LYELLOW_HEX = "FFF9C4"
LPURPLE_HEX = "EDE7F6"
LGREY_HEX   = "F5F5F5"
WHITE_HEX   = "FFFFFF"
DGREEN_HEX  = "1A6B3C"
DRED_HEX    = "C62828"


# ── Indicators ────────────────────────────────────────────────────────────────

def calculate_rsi(close, period=14):
    delta    = close.diff()
    gain     = delta.clip(lower=0)
    loss     = -delta.clip(upper=0)
    avg_gain = gain.rolling(window=period).mean()
    avg_loss = loss.rolling(window=period).mean()
    rs       = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def calculate_macd(close, fast=12, slow=26, signal=9):
    ema_fast    = close.ewm(span=fast,   adjust=False).mean()
    ema_slow    = close.ewm(span=slow,   adjust=False).mean()
    macd_line   = ema_fast - ema_slow
    signal_line = macd_line.ewm(span=signal, adjust=False).mean()
    histogram   = macd_line - signal_line
    return macd_line, signal_line, histogram


def calculate_atr(high, low, close, period=14):
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs()
    ], axis=1).max(axis=1)
    return tr.rolling(window=period).mean()


def calculate_roc(close, period=20):
    return ((close - close.shift(period)) / close.shift(period)) * 100


def calculate_adx(high, low, close, period=14):
    plus_dm  = high.diff()
    minus_dm = low.diff()
    plus_dm  = plus_dm.where((plus_dm > 0) & (plus_dm > minus_dm.abs()), 0)
    minus_dm = minus_dm.abs().where((minus_dm.abs() > 0) & (minus_dm.abs() > plus_dm), 0)
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs()
    ], axis=1).max(axis=1)
    atr_s    = tr.rolling(period).mean()
    plus_di  = 100 * (plus_dm.rolling(period).mean()  / atr_s)
    minus_di = 100 * (minus_dm.rolling(period).mean() / atr_s)
    dx       = (100 * (plus_di - minus_di).abs() / (plus_di + minus_di))
    return dx.rolling(period).mean()


# ── Helpers ───────────────────────────────────────────────────────────────────

def safe(val, multiplier=1, decimals=2, fallback="N/A"):
    try:
        if val is None:
            return fallback
        v = float(val) * multiplier
        return fallback if np.isnan(v) else round(v, decimals)
    except Exception:
        return fallback


def get_market_trend():
    """
    Smart 4-indicator Market Regime.
    Returns tuple: (regime_string, details_dict, score_int)

    Indicators:
      1. SMA50 vs SMA200  — Long term trend (Golden/Death Cross)
      2. Price vs SMA50   — Medium term (is market above/below its trend?)
      3. RSI of Nifty     — Current momentum
      4. % from 52W High  — Depth of correction

    Regimes:
      4/4 → BULLISH
      3/4 → CAUTIOUS BULLISH
      2/4 → CAUTION          ← correctly catches current market
      1/4 → BEARISH
      0/4 → STRONG BEARISH
    """
    try:
        data = yf.download(ETF, period="2y", interval="1d", auto_adjust=True, progress=False)
        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)
        if data.empty or len(data) < 200:
            return "NOT ENOUGH DATA", {}, 0

        close = data["Close"]
        data["SMA50"]    = close.rolling(50).mean()
        data["SMA200"]   = close.rolling(200).mean()
        data["RSI"]      = calculate_rsi(close)
        data["52W_High"] = close.rolling(252, min_periods=200).max()

        latest   = data.iloc[-1]
        price    = float(latest["Close"])
        sma50    = float(latest["SMA50"])
        sma200   = float(latest["SMA200"])
        rsi      = float(latest["RSI"])
        high_52w = float(latest["52W_High"])

        if any(np.isnan(v) for v in [sma50, sma200, rsi, high_52w]):
            return "NOT ENOUGH DATA", {}, 0

        pct_from_high = ((high_52w - price) / high_52w) * 100
        score   = 0
        details = {}

        # 1. SMA50 vs SMA200
        if sma50 > sma200:
            score += 1
            details["SMA Cross"] = f"✅ GOLDEN CROSS (SMA50 {sma50:.0f} > SMA200 {sma200:.0f})"
        else:
            details["SMA Cross"] = f"❌ DEATH CROSS (SMA50 {sma50:.0f} < SMA200 {sma200:.0f})"

        # 2. Price vs SMA50
        if price > sma50:
            score += 1
            details["Price vs SMA50"] = f"✅ Price {price:.0f} ABOVE SMA50 {sma50:.0f}"
        else:
            details["Price vs SMA50"] = f"❌ Price {price:.0f} BELOW SMA50 {sma50:.0f} ({((sma50-price)/sma50*100):.1f}% below)"

        # 3. RSI momentum
        if rsi >= 50:
            score += 1
            details["RSI"] = f"✅ RSI {rsi:.1f} — Bullish momentum"
        else:
            details["RSI"] = f"❌ RSI {rsi:.1f} — Bearish momentum (below 50)"

        # 4. % from 52W High
        if pct_from_high <= 10:
            score += 1
            details["52W High"] = f"✅ Only {pct_from_high:.1f}% below 52W High — Strong zone"
        elif pct_from_high <= 20:
            details["52W High"] = f"⚠️ {pct_from_high:.1f}% below 52W High — Correction territory"
        else:
            details["52W High"] = f"❌ {pct_from_high:.1f}% below 52W High — Deep correction"

        if score == 4:   regime = "BULLISH"
        elif score == 3: regime = "CAUTIOUS BULLISH"
        elif score == 2: regime = "CAUTION"
        elif score == 1: regime = "BEARISH"
        else:            regime = "STRONG BEARISH"

        return regime, details, score

    except Exception as e:
        print(f"Market trend error: {e}")
        return "NOT ENOUGH DATA", {}, 0


# ── Fundamentals ──────────────────────────────────────────────────────────────

def get_fundamentals(stock, mode="swing"):
    empty = {
        "Rev Growth 1Y (%)": "N/A", "Rev Growth 2Y (%)": "N/A",
        "Net Profit Margin (%)": "N/A", "ROE (%)": "N/A", "ROCE (%)": "N/A",
        "EPS (TTM)": "N/A", "Debt/Equity": "N/A", "PE Ratio": "N/A",
        "Fundamental Score": 0, "Fundamental Flags": "Data Error",
    }
    try:
        ticker = yf.Ticker(stock)
        info   = ticker.info

        roe_raw = info.get("returnOnEquity") or info.get("ReturnOnEquity") or None
        roe_pct = float(roe_raw) * 100 if roe_raw is not None else None
        if roe_pct is None:
            try:
                inc = ticker.financials
                bal = ticker.balance_sheet
                if inc is not None and bal is not None and not inc.empty and not bal.empty:
                    net_income = None
                    for label in ["Net Income","NetIncome","Net Income Common Stockholders"]:
                        if label in inc.index:
                            net_income = inc.loc[label].iloc[0]; break
                    equity = None
                    for label in ["Total Stockholder Equity","Stockholders Equity",
                                  "Total Equity Gross Minority Interest","Common Stock Equity"]:
                        if label in bal.index:
                            equity = bal.loc[label].iloc[0]; break
                    if net_income and equity and float(equity) > 0:
                        roe_pct = (float(net_income) / float(equity)) * 100
            except Exception: pass

        pm_raw = info.get("profitMargins") or info.get("netMargins") or None
        pm_pct = float(pm_raw) * 100 if pm_raw is not None else None
        eps    = info.get("trailingEps") or info.get("epsTrailingTwelveMonths") or None
        pe     = info.get("trailingPE")  or info.get("forwardPE") or None

        de_raw = info.get("debtToEquity")
        de = None
        if de_raw is not None:
            de_raw = float(de_raw)
            de = de_raw / 100 if de_raw > 10 else de_raw
        if de is None:
            try:
                bal = ticker.balance_sheet
                if bal is not None and not bal.empty:
                    total_debt = None
                    for label in ["Total Debt","Long Term Debt",
                                  "Total Liabilities Net Minority Interest"]:
                        if label in bal.index:
                            total_debt = bal.loc[label].iloc[0]; break
                    equity = None
                    for label in ["Total Stockholder Equity","Stockholders Equity",
                                  "Common Stock Equity"]:
                        if label in bal.index:
                            equity = bal.loc[label].iloc[0]; break
                    if total_debt is not None and equity and float(equity) > 0:
                        de = float(total_debt) / float(equity)
            except Exception: pass

        roce = None
        try:
            ebit         = info.get("ebitda")
            total_assets = info.get("totalAssets")
            curr_liab    = info.get("totalCurrentLiabilities")
            if ebit and total_assets and curr_liab:
                cap_employed = float(total_assets) - float(curr_liab)
                if cap_employed > 0:
                    roce = (float(ebit) / cap_employed) * 100
            if roce is None:
                bal = ticker.balance_sheet
                inc = ticker.financials
                if bal is not None and inc is not None and not bal.empty and not inc.empty:
                    ebit_val = None
                    for label in ["EBIT","Ebit","Operating Income"]:
                        if label in inc.index:
                            ebit_val = inc.loc[label].iloc[0]; break
                    ta_val = None
                    for label in ["Total Assets"]:
                        if label in bal.index:
                            ta_val = bal.loc[label].iloc[0]; break
                    cl_val = None
                    for label in ["Total Current Liabilities","Current Liabilities"]:
                        if label in bal.index:
                            cl_val = bal.loc[label].iloc[0]; break
                    if ebit_val and ta_val and cl_val:
                        cap_emp = float(ta_val) - float(cl_val)
                        if cap_emp > 0:
                            roce = (float(ebit_val) / cap_emp) * 100
        except Exception: pass

        rev_growth_1y = rev_growth_2y = None
        rg = info.get("revenueGrowth")
        if rg is not None:
            rev_growth_1y = float(rg) * 100
        if rev_growth_1y is None:
            try:
                fin = ticker.financials
                if fin is not None and not fin.empty:
                    rev_row = None
                    for label in ["Total Revenue","Revenue","Net Revenue",
                                  "Operating Revenue","Total Revenues"]:
                        if label in fin.index:
                            rev_row = fin.loc[label]; break
                    if rev_row is not None:
                        rev_sorted = rev_row.sort_index(ascending=False).dropna()
                        if len(rev_sorted) >= 2:
                            r0, r1 = float(rev_sorted.iloc[0]), float(rev_sorted.iloc[1])
                            if r1 != 0: rev_growth_1y = ((r0 - r1) / abs(r1)) * 100
                        if len(rev_sorted) >= 3:
                            r2 = float(rev_sorted.iloc[2])
                            if r2 != 0:
                                rev_growth_2y = ((float(rev_sorted.iloc[0]) - r2) / abs(r2)) * 100
            except Exception: pass

        rev_thresh = 25.0 if mode == "multibagger" else 10.0
        de_thresh  = 0.5  if mode == "multibagger" else 1.0

        fscore = 0; flags = []; skipped = []
        if roe_pct is not None:
            if roe_pct > 15: fscore += 1
            else: flags.append(f"Low ROE ({roe_pct:.1f}%)")
        else: skipped.append("ROE")

        if rev_growth_1y is not None:
            if rev_growth_1y > rev_thresh: fscore += 1
            else: flags.append(f"Slow Growth ({rev_growth_1y:.1f}%<{rev_thresh:.0f}%)")
        else: skipped.append("Rev Growth")

        if pm_pct is not None:
            if pm_pct > 10: fscore += 1
            else: flags.append(f"Low Margin ({pm_pct:.1f}%)")
        else: skipped.append("Margin")

        if de is not None:
            if de < de_thresh: fscore += 1
            else: flags.append(f"High D/E ({de:.2f}>{de_thresh})")
        else: skipped.append("D/E")

        if pe is not None:
            if 0 < float(pe) < 50: fscore += 1
            else: flags.append(f"High PE ({float(pe):.1f})")
        else: skipped.append("PE")

        flag_parts = list(flags)
        if skipped: flag_parts.append(f"Data N/A: {', '.join(skipped)}")
        flag_str = ", ".join(flag_parts) if flag_parts else "All Good"

        return {
            "Rev Growth 1Y (%)":     safe(rev_growth_1y),
            "Rev Growth 2Y (%)":     safe(rev_growth_2y),
            "Net Profit Margin (%)": safe(pm_pct),
            "ROE (%)":               safe(roe_pct),
            "ROCE (%)":              safe(roce),
            "EPS (TTM)":             safe(eps),
            "Debt/Equity":           safe(de),
            "PE Ratio":              safe(pe),
            "Fundamental Score":     fscore,
            "Fundamental Flags":     flag_str,
        }
    except Exception as e:
        print(f"    [{stock}] Fund error: {e}")
        return empty


# ── SWING Screener ────────────────────────────────────────────────────────────

def analyze_swing(stock_list, label=""):
    print(f"  Processing {label} ({len(stock_list)} stocks)...")
    results, skipped = [], []

    for stock in stock_list:
        print(f"    {stock}...", end=" ")
        data = yf.download(stock, period="1y", interval="1d",
                           auto_adjust=True, progress=False)
        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)
        if data.empty or len(data) < 60:
            print("SKIP"); skipped.append(stock); continue

        close  = data["Close"]
        high   = data["High"]
        low    = data["Low"]
        volume = data["Volume"]

        data["SMA20"]    = close.rolling(20).mean()
        data["SMA50"]    = close.rolling(50).mean()
        data["RSI"]      = calculate_rsi(close)
        data["Vol20"]    = volume.rolling(20).mean()
        data["52W_High"] = close.rolling(252, min_periods=200).max()
        data["ATR"]      = calculate_atr(high, low, close)
        data["ROC20"]    = calculate_roc(close)
        data["ADX"]      = calculate_adx(high, low, close)
        macd_line, signal_line, _ = calculate_macd(close)
        data["MACD"]        = macd_line
        data["MACD_Signal"] = signal_line

        latest = data.iloc[-1]
        prev3  = data.iloc[-4] if len(data) >= 4 else data.iloc[-2]
        if any(pd.isna(latest[c]) for c in ["SMA20","SMA50","RSI","MACD","MACD_Signal","ATR"]):
            print("SKIP"); skipped.append(stock); continue

        price = float(latest["Close"])
        atr   = float(latest["ATR"])
        rsi   = float(latest["RSI"])

        above_sma20   = price > float(latest["SMA20"])
        above_sma50   = price > float(latest["SMA50"])
        trend_align   = float(latest["SMA20"]) > float(latest["SMA50"])
        vol_surge     = float(latest["Volume"]) > float(latest["Vol20"]) if not pd.isna(latest["Vol20"]) else False
        high_52w      = float(latest["52W_High"]) if not pd.isna(latest["52W_High"]) else None
        pct_from_high = ((high_52w - price) / high_52w * 100) if high_52w else None
        near_52w      = (pct_from_high <= 10.0) if pct_from_high is not None else False
        macd_now      = float(latest["MACD"])
        sig_now       = float(latest["MACD_Signal"])
        macd_p3       = float(prev3["MACD"])        if not pd.isna(prev3["MACD"])        else macd_now
        sig_p3        = float(prev3["MACD_Signal"]) if not pd.isna(prev3["MACD_Signal"]) else sig_now
        macd_cross    = (macd_now > sig_now) and (macd_p3 <= sig_p3)
        macd_bull     = macd_now > sig_now
        roc_val       = float(latest["ROC20"]) if not pd.isna(latest["ROC20"]) else 0.0
        adx_val       = float(latest["ADX"])   if not pd.isna(latest["ADX"])   else 0.0

        tech_score = sum([
            above_sma20, above_sma50, trend_align,
            rsi > 55, rsi > 60, vol_surge, near_52w,
            macd_cross, roc_val > 5, adx_val > 25,
        ])

        fund     = get_fundamentals(stock, mode="swing")
        combined = tech_score + fund["Fundamental Score"]

        signal = "NO SIGNAL"
        if combined >= 12:  signal = "STRONG SWING"
        elif combined >= 9: signal = "SWING WATCH"

        stop_loss = round(price - 1.5 * atr, 2)
        target    = round(price + 3.0 * atr, 2)
        risk_pct  = round(((price - stop_loss) / price) * 100, 2)

        print(f"Score {combined}/15 | {signal}")
        results.append({
            "Stock":                   stock,
            "Sector":                  SECTOR_MAP.get(stock, "Unknown"),
            "Close":                   round(price, 2),
            "ATR (14)":                round(atr, 2),
            "Stop Loss":               stop_loss,
            "Target (2:1)":            target,
            "Risk (%)":                risk_pct,
            "RSI":                     round(rsi, 2),
            "MACD Crossover":          "YES" if macd_cross else "NO",
            "MACD Bullish":            "YES" if macd_bull  else "NO",
            "Momentum ROC 20D (%)":    round(roc_val, 2),
            "ADX":                     round(adx_val, 2),
            "Volume vs Avg":           "YES" if vol_surge  else "NO",
            "% From 52W High":         safe(pct_from_high),
            "Near 52W High":           "YES" if near_52w   else "NO",
            "Tech Score (max 10)":     tech_score,
            "Rev Growth 1Y (%)":       fund["Rev Growth 1Y (%)"],
            "Rev Growth 2Y (%)":       fund["Rev Growth 2Y (%)"],
            "Net Profit Margin (%)":   fund["Net Profit Margin (%)"],
            "ROE (%)":                 fund["ROE (%)"],
            "ROCE (%)":                fund["ROCE (%)"],
            "EPS (TTM)":               fund["EPS (TTM)"],
            "Debt/Equity":             fund["Debt/Equity"],
            "PE Ratio":                fund["PE Ratio"],
            "Fund Score (max 5)":      fund["Fundamental Score"],
            "Fundamental Flags":       fund["Fundamental Flags"],
            "Combined Score (max 15)": combined,
            "Signal":                  signal,
        })

    if skipped:
        print(f"\n    Skipped: {skipped}")

    df = pd.DataFrame(results)
    df = df.sort_values(
        by=["Combined Score (max 15)", "Tech Score (max 10)", "RSI"],
        ascending=False
    ).reset_index(drop=True)
    df.insert(0, "Rank", df.index + 1)
    return df


# ── MULTIBAGGER Screener ──────────────────────────────────────────────────────

def analyze_multibagger(stock_list):
    print(f"  Processing Smallcap ({len(stock_list)} stocks)...")
    results, skipped = [], []

    for stock in stock_list:
        print(f"    {stock}...", end=" ")
        data = yf.download(stock, period="1y", interval="1d",
                           auto_adjust=True, progress=False)
        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)
        if data.empty or len(data) < 100:
            print("SKIP"); skipped.append(stock); continue

        close  = data["Close"]
        high   = data["High"]
        low    = data["Low"]
        volume = data["Volume"]

        data["SMA50"]    = close.rolling(50).mean()
        data["SMA200"]   = close.rolling(200).mean()
        data["RSI"]      = calculate_rsi(close)
        data["Vol50"]    = volume.rolling(50).mean()
        data["52W_High"] = close.rolling(252, min_periods=200).max()
        data["52W_Low"]  = close.rolling(252, min_periods=200).min()
        data["ATR"]      = calculate_atr(high, low, close)
        data["ROC20"]    = calculate_roc(close)
        data["ADX"]      = calculate_adx(high, low, close)

        latest = data.iloc[-1]
        if any(pd.isna(latest[c]) for c in ["SMA50","SMA200","RSI","ATR"]):
            print("SKIP"); skipped.append(stock); continue

        price    = float(latest["Close"])
        atr      = float(latest["ATR"])
        rsi      = float(latest["RSI"])
        adx_val  = float(latest["ADX"])   if not pd.isna(latest["ADX"])   else 0.0
        roc_val  = float(latest["ROC20"]) if not pd.isna(latest["ROC20"]) else 0.0
        sma50    = float(latest["SMA50"])
        sma200   = float(latest["SMA200"])
        vol50    = float(latest["Vol50"]) if not pd.isna(latest["Vol50"]) else 0.0
        high_52w = float(latest["52W_High"]) if not pd.isna(latest["52W_High"]) else None
        low_52w  = float(latest["52W_Low"])  if not pd.isna(latest["52W_Low"])  else None

        pct_from_high = ((high_52w - price) / high_52w * 100) if high_52w else None
        pct_from_low  = ((price - low_52w)  / low_52w  * 100) if low_52w  else None

        above_sma200   = price > sma200
        golden_cross   = sma50 > sma200
        near_52w_break = (pct_from_high <= 5.0) if pct_from_high is not None else False
        vol_expansion  = float(latest["Volume"]) > (vol50 * 1.5) if vol50 > 0 else False
        rsi_breakout   = 55 <= rsi <= 75
        strong_adx     = adx_val > 30
        strong_roc     = roc_val > 10.0

        tech_score = sum([
            above_sma200, golden_cross, near_52w_break,
            vol_expansion, rsi_breakout, strong_adx, strong_roc,
        ])

        fund     = get_fundamentals(stock, mode="multibagger")
        combined = tech_score + fund["Fundamental Score"]

        signal = "NO SIGNAL"
        if combined >= 9:   signal = "MULTIBAGGER SETUP"
        elif combined >= 7: signal = "BREAKOUT WATCH"

        stop_loss = round(price - 3.0 * atr, 2)
        target_15 = round(price * 1.15, 2)
        target_20 = round(price * 1.20, 2)
        risk_pct  = round(((price - stop_loss) / price) * 100, 2)

        print(f"Score {combined}/12 | {signal}")
        results.append({
            "Stock":                   stock,
            "Sector":                  SECTOR_MAP.get(stock, "Unknown"),
            "Close":                   round(price, 2),
            "ATR (14)":                round(atr, 2),
            "Stop Loss (3x ATR)":      stop_loss,
            "Target 15%":              target_15,
            "Target 20%":              target_20,
            "Risk (%)":                risk_pct,
            "RSI":                     round(rsi, 2),
            "ADX":                     round(adx_val, 2),
            "Momentum ROC 20D (%)":    round(roc_val, 2),
            "Volume Expansion":        "YES" if vol_expansion  else "NO",
            "Above SMA200":            "YES" if above_sma200   else "NO",
            "Golden Cross":            "YES" if golden_cross   else "NO",
            "Near 52W High (<5%)":     "YES" if near_52w_break else "NO",
            "% From 52W High":         safe(pct_from_high),
            "% From 52W Low":          safe(pct_from_low),
            "Tech Score (max 7)":      tech_score,
            "Rev Growth 1Y (%)":       fund["Rev Growth 1Y (%)"],
            "Net Profit Margin (%)":   fund["Net Profit Margin (%)"],
            "ROE (%)":                 fund["ROE (%)"],
            "Debt/Equity":             fund["Debt/Equity"],
            "PE Ratio":                fund["PE Ratio"],
            "Fund Score (max 5)":      fund["Fundamental Score"],
            "Fundamental Flags":       fund["Fundamental Flags"],
            "Combined Score (max 12)": combined,
            "Signal":                  signal,
        })

    if skipped:
        print(f"\n    Skipped: {len(skipped)} stocks")

    df = pd.DataFrame(results)
    df = df.sort_values(
        by=["Combined Score (max 12)", "Tech Score (max 7)", "RSI"],
        ascending=False
    ).reset_index(drop=True)
    df.insert(0, "Rank", df.index + 1)
    return df


# ── Excel Style Helpers ───────────────────────────────────────────────────────

def mf(hex_color):  return PatternFill("solid", fgColor=hex_color)
def mfont(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")
def mborder():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)
def mcenter(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def mleft():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Format Swing Sheet ────────────────────────────────────────────────────────

def format_swing_sheet(ws, header_color=NAVY_HEX):
    col_map = {cell.value: cell.column for cell in ws[1]}

    for cell in ws[1]:
        cell.fill = mf(header_color); cell.font = mfont(bold=True, color=WHITE_HEX, size=9)
        cell.alignment = mcenter(); cell.border = mborder()
    ws.row_dimensions[1].height = 38

    widths = {
        "Rank": 5, "Stock": 13, "Sector": 13, "Close": 9, "ATR (14)": 9,
        "Stop Loss": 10, "Target (2:1)": 10, "Risk (%)": 8, "RSI": 7,
        "MACD Crossover": 10, "MACD Bullish": 9, "Momentum ROC 20D (%)": 12,
        "ADX": 7, "Volume vs Avg": 10, "% From 52W High": 12, "Near 52W High": 10,
        "Tech Score (max 10)": 11, "Rev Growth 1Y (%)": 11, "Rev Growth 2Y (%)": 11,
        "Net Profit Margin (%)": 12, "ROE (%)": 8, "ROCE (%)": 8, "EPS (TTM)": 8,
        "Debt/Equity": 9, "PE Ratio": 8, "Fund Score (max 5)": 10,
        "Fundamental Flags": 22, "Combined Score (max 15)": 13, "Signal": 13,
    }
    for col, w in widths.items():
        if col in col_map:
            ws.column_dimensions[get_column_letter(col_map[col])].width = w

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        row_fill = mf(LGREY_HEX) if row_idx % 2 == 0 else mf(WHITE_HEX)
        for cell in row:
            cell.border = mborder(); cell.alignment = mcenter()
            cell.font = mfont(size=9); cell.fill = row_fill

        def col(name): return col_map.get(name)

        if col("Signal"):
            c = row[col("Signal") - 1]
            if c.value == "STRONG SWING":
                c.fill = mf("1A6B3C"); c.font = mfont(bold=True, color=WHITE_HEX, size=9)
            elif c.value == "SWING WATCH":
                c.fill = mf("E65100"); c.font = mfont(bold=True, color=WHITE_HEX, size=9)
            else:
                c.fill = mf("757575"); c.font = mfont(bold=True, color=WHITE_HEX, size=9)

        if col("Combined Score (max 15)"):
            c = row[col("Combined Score (max 15)") - 1]
            try:
                v = int(c.value)
                if v >= 12:  c.fill = mf(LGREEN_HEX); c.font = mfont(bold=True, size=9)
                elif v >= 9: c.fill = mf(LYELLOW_HEX)
                else:        c.fill = mf(LRED_HEX)
            except Exception: pass

        if col("RSI"):
            c = row[col("RSI") - 1]
            try:
                v = float(c.value)
                if v >= 70:   c.fill = mf(LRED_HEX);   c.font = mfont(bold=True, size=9)
                elif v >= 55: c.fill = mf(LGREEN_HEX)
                elif v < 40:  c.fill = mf(LRED_HEX)
            except Exception: pass

        for yn in ["Volume vs Avg", "Near 52W High", "MACD Crossover", "MACD Bullish"]:
            if col(yn):
                c = row[col(yn) - 1]
                if c.value == "YES":
                    c.fill = mf(LGREEN_HEX); c.font = mfont(bold=True, size=9, color=DGREEN_HEX)
                elif c.value == "NO":
                    c.fill = mf(LRED_HEX); c.font = mfont(color=DRED_HEX, size=9)

        if col("Momentum ROC 20D (%)"):
            c = row[col("Momentum ROC 20D (%)") - 1]
            try:
                v = float(c.value)
                if v > 5:   c.fill = mf(LGREEN_HEX)
                elif v < 0: c.fill = mf(LRED_HEX)
            except Exception: pass

        if col("ADX"):
            c = row[col("ADX") - 1]
            try:
                if float(c.value) >= 25:
                    c.fill = mf(LGREEN_HEX); c.font = mfont(bold=True, size=9)
            except Exception: pass

        if col("Stop Loss"):
            c = row[col("Stop Loss") - 1]
            c.fill = mf("FFF3E0"); c.font = mfont(bold=True, color="E65100", size=9)
        if col("Target (2:1)"):
            c = row[col("Target (2:1)") - 1]
            c.fill = mf(LGREEN_HEX); c.font = mfont(bold=True, color=DGREEN_HEX, size=9)

        if col("Fundamental Flags"):
            c = row[col("Fundamental Flags") - 1]
            c.alignment = mleft()
            if c.value == "All Good":
                c.fill = mf(LGREEN_HEX); c.font = mfont(bold=True, size=9, color=DGREEN_HEX)

        if col("Debt/Equity"):
            c = row[col("Debt/Equity") - 1]
            try:
                if float(c.value) > 1.0:
                    c.fill = mf(LRED_HEX); c.font = mfont(color=DRED_HEX, size=9)
            except Exception: pass

    ws.freeze_panes = "A2"


# ── Format Multibagger Sheet ──────────────────────────────────────────────────

def format_multibagger_sheet(ws):
    col_map = {cell.value: cell.column for cell in ws[1]}

    for cell in ws[1]:
        cell.fill = mf(PURPLE_HEX); cell.font = mfont(bold=True, color=WHITE_HEX, size=9)
        cell.alignment = mcenter(); cell.border = mborder()
    ws.row_dimensions[1].height = 38

    widths = {
        "Rank": 5, "Stock": 13, "Sector": 13, "Close": 9, "ATR (14)": 9,
        "Stop Loss (3x ATR)": 13, "Target 15%": 10, "Target 20%": 10, "Risk (%)": 8,
        "RSI": 7, "ADX": 7, "Momentum ROC 20D (%)": 12,
        "Volume Expansion": 11, "Above SMA200": 10, "Golden Cross": 10,
        "Near 52W High (<5%)": 13, "% From 52W High": 12, "% From 52W Low": 12,
        "Tech Score (max 7)": 11, "Rev Growth 1Y (%)": 11,
        "Net Profit Margin (%)": 12, "ROE (%)": 8, "Debt/Equity": 9, "PE Ratio": 8,
        "Fund Score (max 5)": 10, "Fundamental Flags": 24,
        "Combined Score (max 12)": 13, "Signal": 16,
    }
    for col, w in widths.items():
        if col in col_map:
            ws.column_dimensions[get_column_letter(col_map[col])].width = w

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        row_fill = mf(LGREY_HEX) if row_idx % 2 == 0 else mf(WHITE_HEX)
        for cell in row:
            cell.border = mborder(); cell.alignment = mcenter()
            cell.font = mfont(size=9); cell.fill = row_fill

        def col(name): return col_map.get(name)

        if col("Signal"):
            c = row[col("Signal") - 1]
            if c.value == "MULTIBAGGER SETUP":
                c.fill = mf("4A0E8F"); c.font = mfont(bold=True, color=WHITE_HEX, size=9)
            elif c.value == "BREAKOUT WATCH":
                c.fill = mf("7B1FA2"); c.font = mfont(bold=True, color=WHITE_HEX, size=9)
            else:
                c.fill = mf("757575"); c.font = mfont(bold=True, color=WHITE_HEX, size=9)

        if col("Combined Score (max 12)"):
            c = row[col("Combined Score (max 12)") - 1]
            try:
                v = int(c.value)
                if v >= 9:   c.fill = mf(LPURPLE_HEX); c.font = mfont(bold=True, size=9)
                elif v >= 7: c.fill = mf(LYELLOW_HEX)
                else:        c.fill = mf(LRED_HEX)
            except Exception: pass

        if col("RSI"):
            c = row[col("RSI") - 1]
            try:
                v = float(c.value)
                if 55 <= v <= 75: c.fill = mf(LPURPLE_HEX)
                elif v > 75:      c.fill = mf(LRED_HEX)
            except Exception: pass

        if col("ADX"):
            c = row[col("ADX") - 1]
            try:
                if float(c.value) >= 30:
                    c.fill = mf(LPURPLE_HEX); c.font = mfont(bold=True, size=9)
            except Exception: pass

        if col("Momentum ROC 20D (%)"):
            c = row[col("Momentum ROC 20D (%)") - 1]
            try:
                v = float(c.value)
                if v > 10:  c.fill = mf(LPURPLE_HEX)
                elif v < 0: c.fill = mf(LRED_HEX)
            except Exception: pass

        for yn in ["Volume Expansion","Above SMA200","Golden Cross","Near 52W High (<5%)"]:
            if col(yn):
                c = row[col(yn) - 1]
                if c.value == "YES":
                    c.fill = mf(LPURPLE_HEX); c.font = mfont(bold=True, size=9, color="4A0E8F")
                elif c.value == "NO":
                    c.fill = mf(LRED_HEX); c.font = mfont(color=DRED_HEX, size=9)

        if col("Stop Loss (3x ATR)"):
            c = row[col("Stop Loss (3x ATR)") - 1]
            c.fill = mf("FFF3E0"); c.font = mfont(bold=True, color="E65100", size=9)

        for tgt in ["Target 15%", "Target 20%"]:
            if col(tgt):
                c = row[col(tgt) - 1]
                c.fill = mf(LPURPLE_HEX); c.font = mfont(bold=True, color="4A0E8F", size=9)

        if col("Fundamental Flags"):
            c = row[col("Fundamental Flags") - 1]
            c.alignment = mleft()
            if c.value == "All Good":
                c.fill = mf(LPURPLE_HEX); c.font = mfont(bold=True, size=9, color="4A0E8F")

        if col("Debt/Equity"):
            c = row[col("Debt/Equity") - 1]
            try:
                if float(c.value) > 0.5:
                    c.fill = mf(LRED_HEX); c.font = mfont(color=DRED_HEX, size=9)
            except Exception: pass

    ws.freeze_panes = "A2"


# ── Format Market Regime Sheet (SMART 4-INDICATOR) ───────────────────────────

def format_market_regime_sheet(wb, market_trend_tuple):
    if "Market Regime" in wb.sheetnames:
        del wb["Market Regime"]
    ws2 = wb.create_sheet("Market Regime")

    if isinstance(market_trend_tuple, tuple):
        regime, details, score = market_trend_tuple
    else:
        regime, details, score = market_trend_tuple, {}, 0

    regime_config = {
        "BULLISH":          ("1A6B3C", "🟢 BULLISH"),
        "CAUTIOUS BULLISH": ("2E7D32", "🟡 CAUTIOUS BULLISH"),
        "CAUTION":          ("E65100", "🟠 CAUTION"),
        "BEARISH":          ("C62828", "🔴 BEARISH"),
        "STRONG BEARISH":   ("880E4F", "⛔ STRONG BEARISH"),
        "NOT ENOUGH DATA":  ("757575", "⚪ NOT ENOUGH DATA"),
    }
    bg_color, icon = regime_config.get(regime, ("757575", regime))

    action_messages = {
        "BULLISH":          ("All 4 indicators bullish. Act on STRONG SWING signals with normal sizes. Trail stops.",
                             "✅ Full position sizes\n✅ Act on STRONG SWING signals\n✅ Trail stop losses"),
        "CAUTIOUS BULLISH": ("3/4 indicators bullish. Long term trend intact but some weakness. Be selective.",
                             "✅ Reduce positions to 75%\n✅ Only STRONG SWING score 12+\n⚠️ Keep stop losses tight"),
        "CAUTION":          ("2/4 indicators bullish. Market showing clear weakness. SMA cross may be bullish but price and RSI are weak.",
                             "⚠️ Reduce positions to 50%\n⚠️ Only top 3 ranked stocks\n⚠️ Consider waiting for improvement"),
        "BEARISH":          ("1/4 indicators bullish. Clear downtrend. Avoid new long positions. Protect capital.",
                             "🛑 No new long positions\n🛑 Exit weak holdings\n⚠️ Stay in cash until regime improves"),
        "STRONG BEARISH":   ("All 4 indicators bearish. Very high risk. Stay in cash.",
                             "🛑 Stay in cash\n🛑 No new positions\n🛑 Wait for CAUTION regime before re-entering"),
    }
    message, action_text = action_messages.get(regime, ("Insufficient data.", "No action."))

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 68
    for r, h in {1:24, 2:50, 3:40, 4:60, 5:22, 6:22, 7:22, 8:22, 9:22}.items():
        ws2.row_dimensions[r].height = h

    ws2["A1"] = "MARKET REGIME"
    ws2["A1"].fill = mf(NAVY_HEX); ws2["A1"].font = mfont(bold=True, color=WHITE_HEX, size=12)
    ws2["A1"].alignment = mcenter(); ws2.merge_cells("A1:B1")

    ws2["A2"] = "Current Regime"
    ws2["B2"] = f"{icon}   (Score: {score}/4)"
    ws2["A2"].fill = mf(LGREY_HEX); ws2["B2"].fill = mf(bg_color)
    ws2["A2"].font = mfont(bold=True, size=11)
    ws2["B2"].font = Font(bold=True, color=WHITE_HEX, size=16, name="Arial")
    ws2["A2"].alignment = ws2["B2"].alignment = mcenter()

    ws2["A3"] = "What This Means"
    ws2["B3"] = message
    ws2["A3"].font = mfont(bold=True, size=10); ws2["B3"].font = mfont(size=10)
    ws2["B3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws2["A3"].alignment = mcenter()

    ws2["A4"] = "Action Guide"
    ws2["B4"] = action_text
    ws2["A4"].font = mfont(bold=True, size=10); ws2["B4"].font = mfont(size=10)
    ws2["B4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws2["A4"].alignment = mcenter()

    ws2["A5"] = "INDICATOR BREAKDOWN"
    ws2["B5"] = "Status"
    ws2["A5"].fill = ws2["B5"].fill = mf(TEAL_HEX)
    ws2["A5"].font = ws2["B5"].font = mfont(bold=True, color=WHITE_HEX, size=10)
    ws2["A5"].alignment = ws2["B5"].alignment = mcenter()

    for i, label in enumerate(["SMA Cross", "Price vs SMA50", "RSI", "52W High"]):
        row = 6 + i
        detail = details.get(label, "N/A")
        ws2[f"A{row}"] = label
        ws2[f"B{row}"] = detail
        ws2[f"A{row}"].font = mfont(bold=True, size=10)
        ws2[f"B{row}"].font = mfont(size=10)
        ws2[f"A{row}"].alignment = mcenter()
        ws2[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if detail.startswith("✅"):
            ws2[f"A{row}"].fill = ws2[f"B{row}"].fill = mf(LGREEN_HEX)
        elif detail.startswith("❌"):
            ws2[f"A{row}"].fill = ws2[f"B{row}"].fill = mf(LRED_HEX)
        else:
            ws2[f"A{row}"].fill = ws2[f"B{row}"].fill = mf(LYELLOW_HEX)

    for row in ws2.iter_rows(min_row=2, max_row=9, min_col=1, max_col=2):
        for cell in row:
            cell.border = mborder()


# ── Format Full Excel ─────────────────────────────────────────────────────────

def format_excel(filepath, market_trend_tuple):
    wb = load_workbook(filepath)

    if "Nifty 50 Signals"  in wb.sheetnames: format_swing_sheet(wb["Nifty 50 Signals"],  NAVY_HEX)
    if "Next 50 Signals"   in wb.sheetnames: format_swing_sheet(wb["Next 50 Signals"],   TEAL_HEX)
    if "Smallcap Breakout" in wb.sheetnames: format_multibagger_sheet(wb["Smallcap Breakout"])

    format_market_regime_sheet(wb, market_trend_tuple)

    # Score Legend
    if "Score Legend" in wb.sheetnames: del wb["Score Legend"]
    ws3 = wb.create_sheet("Score Legend")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 46

    legend = [
        ("SWING SIGNALS — Nifty 50 & Next 50", "", "Combined max 15"),
        ("Signal",               "Score",    "Meaning"),
        ("STRONG SWING",         "12-15",    "High conviction. Research and enter."),
        ("SWING WATCH",          "9-11",     "Forming up. Monitor for confirmation."),
        ("NO SIGNAL",            "0-8",      "Does not qualify today. Skip."),
        ("", "", ""),
        ("TECH SCORE (max 10)", "", ""),
        ("Above SMA20",          "+1", "Price above 20-day moving average"),
        ("Above SMA50",          "+1", "Price above 50-day moving average"),
        ("SMA20 > SMA50",        "+1", "Short-term trend above long-term"),
        ("RSI > 55",             "+1", "Bullish momentum"),
        ("RSI > 60",             "+1", "Strong momentum"),
        ("Volume > 20D Avg",     "+1", "Institutional participation"),
        ("Within 10% of 52W High","+1","Breakout zone"),
        ("MACD Crossover",       "+1", "MACD crossed above signal (3-day window)"),
        ("ROC 20D > 5%",         "+1", "Strong price momentum"),
        ("ADX > 25",             "+1", "Trend is strong, not sideways"),
        ("", "", ""),
        ("FUND SCORE (max 5)", "", "Swing thresholds"),
        ("ROE > 15%",            "+1", "Good return on equity"),
        ("Rev Growth > 10%",     "+1", "Business expanding"),
        ("Net Margin > 10%",     "+1", "Profitable business"),
        ("D/E < 1.0",            "+1", "Not over-leveraged"),
        ("PE < 50",              "+1", "Not wildly overvalued"),
        ("", "", ""),
        ("MULTIBAGGER SIGNALS — Smallcap", "", "Combined max 12"),
        ("Signal",               "Score",    "Meaning"),
        ("MULTIBAGGER SETUP",    "9-12",     "High conviction breakout with strong fundamentals"),
        ("BREAKOUT WATCH",       "7-8",      "Technical breakout forming. Watch closely."),
        ("NO SIGNAL",            "0-6",      "Not ready. Skip."),
        ("", "", ""),
        ("TECH SCORE (max 7)", "", "Stricter than swing"),
        ("Above SMA200",         "+1", "Long-term uptrend confirmed"),
        ("SMA50 > SMA200",       "+1", "Golden Cross — major bullish signal"),
        ("Within 5% of 52W High","+1", "Near breakout point"),
        ("Volume > 1.5x 50D Avg","+1", "Strong institutional accumulation"),
        ("RSI 55-75",            "+1", "Momentum — not yet overbought"),
        ("ADX > 30",             "+1", "Very strong trend forming"),
        ("ROC 20D > 10%",        "+1", "Strong multi-week momentum"),
        ("", "", ""),
        ("FUND SCORE (max 5)", "", "Stricter multibagger thresholds"),
        ("ROE > 15%",            "+1", "Good return on equity"),
        ("Rev Growth > 25%",     "+1", "HIGH GROWTH required for multibaggers"),
        ("Net Margin > 10%",     "+1", "Profitable business"),
        ("D/E < 0.5",            "+1", "CLEAN balance sheet — low debt"),
        ("PE < 50",              "+1", "Reasonable valuation"),
        ("", "", ""),
        ("MARKET REGIME (4-indicator)", "", ""),
        ("BULLISH",              "4/4", "All clear. Full position sizes."),
        ("CAUTIOUS BULLISH",     "3/4", "Mostly good. Reduce to 75%."),
        ("CAUTION",              "2/4", "Weakness present. Reduce to 50%."),
        ("BEARISH",              "1/4", "Downtrend. No new longs."),
        ("STRONG BEARISH",       "0/4", "Stay in cash completely."),
        ("", "", ""),
        ("ATR TRADE SIZING", "", ""),
        ("Swing Stop",           "1.5x ATR",  "Below entry price"),
        ("Swing Target",         "3.0x ATR",  "2:1 risk-reward"),
        ("Smallcap Stop",        "3.0x ATR",  "Wider — volatile stocks need room"),
        ("Smallcap Target 1",    "+15%",       "First target — take partial profits here"),
        ("Smallcap Target 2",    "+20%",       "Stretch — trail stop after Target 1"),
    ]

    hdr_rows = {0, 6, 17, 24, 30, 38, 45, 51, 57}
    sub_rows = {1, 26}
    for r_idx, (a, b, c_val) in enumerate(legend, start=1):
        ws3.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate([a, b, c_val], start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = mborder()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = mfont(size=9)
            if r_idx - 1 in hdr_rows:
                cell.fill = mf(NAVY_HEX); cell.font = mfont(bold=True, color=WHITE_HEX, size=9)
            elif r_idx - 1 in sub_rows:
                cell.fill = mf(TEAL_HEX); cell.font = mfont(bold=True, color=WHITE_HEX, size=9)
            elif r_idx % 2 == 0:
                cell.fill = mf(LGREY_HEX)
            if a == "STRONG SWING":        cell.fill = mf(LGREEN_HEX)
            if a == "SWING WATCH":         cell.fill = mf(LYELLOW_HEX)
            if a == "MULTIBAGGER SETUP":   cell.fill = mf(LPURPLE_HEX)
            if a == "BREAKOUT WATCH":      cell.fill = mf(LYELLOW_HEX)
            if a == "BULLISH":             cell.fill = mf(LGREEN_HEX)
            if a == "CAUTIOUS BULLISH":    cell.fill = mf("DCEDC8")
            if a == "CAUTION":             cell.fill = mf(LYELLOW_HEX)
            if a == "BEARISH":             cell.fill = mf(LRED_HEX)
            if a == "STRONG BEARISH":      cell.fill = mf("F8BBD0")

    wb.save(filepath)
    print("Excel formatting applied.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main(send_alerts=True):
    import os
    os.makedirs("output", exist_ok=True)

    print("=" * 60)
    print("  TRADING SYSTEM — DAILY SCAN")
    print("=" * 60)

    print("\nChecking Market Trend...")
    market_trend_tuple = get_market_trend()
    if isinstance(market_trend_tuple, tuple):
        regime, details, score = market_trend_tuple
    else:
        regime, details, score = market_trend_tuple, {}, 0
    print(f"Market Regime: {regime} (Score: {score}/4)")
    for k, v in details.items():
        print(f"  {k}: {v}")

    print("\n[1/3] Nifty 50 Swing Screener")
    nifty50_df = analyze_swing(STOCKS, label="Nifty 50")

    print("\n[2/3] Nifty Next 50 Swing Screener")
    next50_df = analyze_swing(STOCKS_NEXT50, label="Next 50")

    print("\n[3/3] Smallcap Multibagger Screener")
    smallcap_df = analyze_multibagger(STOCKS_SMALLCAP)

    print("\nWriting Excel...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        nifty50_df.to_excel(writer,  sheet_name="Nifty 50 Signals",  index=False)
        next50_df.to_excel(writer,   sheet_name="Next 50 Signals",   index=False)
        smallcap_df.to_excel(writer, sheet_name="Smallcap Breakout", index=False)
        pd.DataFrame({"Market Regime": [regime], "Score": [f"{score}/4"]}).to_excel(
            writer, sheet_name="Market Regime", index=False)

    print("Applying formatting...")
    format_excel(OUTPUT_FILE, market_trend_tuple)

    print(f"\n{'='*60}")
    print(f"  SCAN COMPLETE  |  Market: {regime} ({score}/4)")
    print(f"{'='*60}")
    for label, df, s1, s2 in [
        ("Nifty 50",  nifty50_df,  "STRONG SWING",      "SWING WATCH"),
        ("Next 50",   next50_df,   "STRONG SWING",      "SWING WATCH"),
        ("Smallcap",  smallcap_df, "MULTIBAGGER SETUP", "BREAKOUT WATCH"),
    ]:
        c1 = df[df["Signal"] == s1]
        c2 = df[df["Signal"] == s2]
        sc_col = [c for c in df.columns if "Combined Score" in c]
        sc_col = sc_col[0] if sc_col else "RSI"
        print(f"\n  {label} ({len(df)} stocks):")
        print(f"    {s1}: {len(c1)} | {s2}: {len(c2)}")
        if not c1.empty:
            for _, r in c1.head(3).iterrows():
                print(f"      #{int(r['Rank'])} {r['Stock']} | Score {r[sc_col]} | RSI {r['RSI']}")

    print(f"\n  Saved: {OUTPUT_FILE}")
    print(f"{'='*60}")

    return nifty50_df, next50_df, smallcap_df, regime


if __name__ == "__main__":
    main()
