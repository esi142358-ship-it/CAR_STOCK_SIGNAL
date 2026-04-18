"""
SMA Filter — USA Stocks
================================
Checks every stock against 3 SMA conditions:
  ✅ Price > 50 SMA
  ✅ Price > 100 SMA
  ✅ Price > 200 SMA
  ✅ Price NOT more than 10% above 200 SMA
     (sweet spot — not overbought)

Reads:   us_master_reference.csv
         us_buy_signals_report.xlsx (today's CAR signals)
Saves:   us_us_sma_scores.csv          (reusable — no re-download)
Creates: us_us_sma_filter_report.xlsx

Sheets:
  1. All Stocks SMA      — all stocks with SMA values + % diff
  2. SMA Sweet Spot      — only stocks meeting ALL 4 conditions
  3. Below 200 SMA       — stocks in downtrend (avoid)
  4. Overbought >10%     — stocks too far above 200 SMA
  5. CAR + SMA Combined  — CAR signals that also pass SMA filter ⭐
  6. How to Use

Usage:
    python us_sma_filter.py

Tip: Run AFTER car_signal_scanner.py for combined report!
"""

import yfinance as yf
import pandas as pd
import os, time, warnings, tempfile
from datetime import date, timedelta
warnings.filterwarnings("ignore")

_cache_dir = tempfile.mkdtemp(prefix="yf_sma_")
try:
    yf.set_tz_cache_location(_cache_dir)
except Exception:
    pass

# ── FILES ─────────────────────────────────────────────────────────────────────
MASTER_FILE  = "us_master_reference.csv"
SIGNALS_FILE = "us_buy_signals_report.xlsx"
SCORES_CSV   = "us_sma_scores.csv"
OUTPUT_FILE  = "us_sma_filter_report.xlsx"
DELAY        = 0.5
MAX_RETRIES  = 3

# ── SMA CONDITION ─────────────────────────────────────────────────────────────
SMA_ABOVE_PCT  = 0.0    # price must be ABOVE 200 SMA (0%)
SMA_MAX_PCT    = 10.0   # price must NOT be more than 10% above 200 SMA
# ─────────────────────────────────────────────────────────────────────────────


def nse_ticker(symbol):  # USA — no suffix needed
    return symbol  # USA stocks — no .NS suffix


def fetch_sma(symbol):
    """Fetch 50, 100, 200 SMA for a stock."""
    ticker = nse_ticker(symbol)
    end    = date.today()
    start  = end - timedelta(days=420)   # need 200+ trading days

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            t    = yf.Ticker(ticker)
            hist = t.history(start=str(start), end=str(end),
                             auto_adjust=True, actions=False)

            if hist is None or hist.empty or len(hist) < 50:
                return None

            # Flatten MultiIndex if present
            if isinstance(hist.columns, pd.MultiIndex):
                hist.columns = hist.columns.get_level_values(0)

            closes = hist["Close"].dropna()
            if len(closes) < 50:
                return None

            curr_price = round(float(closes.iloc[-1]), 2)
            sma50  = round(float(closes.tail(50).mean()),  2) if len(closes) >= 50  else None
            sma100 = round(float(closes.tail(100).mean()), 2) if len(closes) >= 100 else None
            sma200 = round(float(closes.tail(200).mean()), 2) if len(closes) >= 200 else None

            # % difference from each SMA
            def pct_diff(price, sma):
                if sma and sma > 0:
                    return round((price - sma) / sma * 100, 2)
                return None

            diff50  = pct_diff(curr_price, sma50)
            diff100 = pct_diff(curr_price, sma100)
            diff200 = pct_diff(curr_price, sma200)

            # Conditions
            above50  = diff50  is not None and diff50  > 0
            above100 = diff100 is not None and diff100 > 0
            above200 = diff200 is not None and diff200 > 0
            not_overbought = diff200 is not None and diff200 <= SMA_MAX_PCT

            # Sweet spot = above all 3 AND within 10% of 200 SMA
            sweet_spot = above50 and above100 and above200 and not_overbought

            # Signal strength label
            if sweet_spot:
                if diff200 <= 5:
                    signal = "STRONG ⭐⭐⭐"
                else:
                    signal = "GOOD ⭐⭐"
            elif above200 and not not_overbought:
                signal = "OVERBOUGHT ⚠️"
            elif above200:
                signal = "ABOVE 200 ✅"
            elif above100:
                signal = "ABOVE 100"
            elif above50:
                signal = "ABOVE 50 ONLY"
            else:
                signal = "BELOW ALL ❌"

            return {
                "Price":         curr_price,
                "SMA 50":        sma50,
                "SMA 100":       sma100,
                "SMA 200":       sma200,
                "% vs SMA 50":   diff50,
                "% vs SMA 100":  diff100,
                "% vs SMA 200":  diff200,
                "Above 50":      "✅" if above50  else "❌",
                "Above 100":     "✅" if above100 else "❌",
                "Above 200":     "✅" if above200 else "❌",
                "Within 10%":    "✅" if not_overbought and above200 else "❌",
                "Sweet Spot":    "✅ YES" if sweet_spot else "❌ NO",
                "SMA Signal":    signal,
                "Data Points":   len(closes),
            }

        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(1)
            else:
                return None
    return None


def build_sma_scores(master_df):
    """Fetch SMA data for all stocks."""
    total   = len(master_df)
    records = []

    print(f"\nFetching SMA data for {total} stocks...")
    print("(~10-15 min for 500 stocks)\n")

    for i, row in enumerate(master_df.itertuples(index=False), 1):
        rd       = row._asdict()
        symbol   = str(rd.get("Symbol", "")).strip().upper()
        cap      = rd.get("Cap Type", "")
        sector   = rd.get("Sector", "")
        idx_cat  = rd.get("Index Category", "")
        company  = rd.get("Company Name", symbol)

        print(f"  [{i:>3}/{total}] {symbol:<12}", end="", flush=True)

        sma = fetch_sma(symbol)

        if sma:
            print(f" ₹{sma['Price']:>8}  "
                  f"SMA200:{sma['SMA 200']:>8}  "
                  f"Diff:{sma['% vs SMA 200']:>+6.1f}%  "
                  f"{sma['SMA Signal']}")
        else:
            print(" no data")
            sma = {
                "Price": None, "SMA 50": None,
                "SMA 100": None, "SMA 200": None,
                "% vs SMA 50": None, "% vs SMA 100": None,
                "% vs SMA 200": None,
                "Above 50": "?", "Above 100": "?",
                "Above 200": "?", "Within 10%": "?",
                "Sweet Spot": "❓ NO DATA",
                "SMA Signal": "NO DATA ❓",
                "Data Points": 0,
            }

        record = {
            "Symbol":         symbol,
            "Company Name":   company,
            "Index Category": idx_cat,
            "Cap Type":       cap,
            "Sector":         sector,
            **sma,
            "Scored On":      str(date.today()),
        }
        records.append(record)
        time.sleep(DELAY)

    return pd.DataFrame(records)


def write_excel(scores_df, signals_df=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb        = Workbook()
    DARK_BLUE = "1F4E79"; MID_BLUE  = "2E75B6"; TEAL    = "1F7391"
    WHITE     = "FFFFFF"; LGREY     = "F5F5F5"
    GREEN_BG  = "E2EFDA"; RED_BG    = "FFE2E2"
    GOLD      = "FFF2CC"; ORANGE    = "FCE4D6"
    GREEN_C   = "375623"; RED_C     = "C00000"
    MOD_C     = "7F6000"; SOFT_C    = "833C00"
    STAR3_BG  = "C6EFCE"; STAR2_BG  = "FFEB9C"

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft    = Alignment(horizontal="left",   vertical="center")

    def hdr(cell, bg=DARK_BLUE, fg=WHITE):
        cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = ctr; cell.border = border

    def dat(cell, bg=None, bold=False, color="000000", align=None):
        cell.font      = Font(bold=bold, color=color, name="Arial", size=10)
        cell.alignment = align if align else ctr
        cell.border    = border
        if bg: cell.fill = PatternFill("solid", fgColor=bg)

    def add_title(ws, text, ncols, bg=DARK_BLUE, fg=WHITE):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        tc = ws["A1"]
        tc.value     = text
        tc.font      = Font(bold=True, color=fg, name="Arial", size=13)
        tc.fill      = PatternFill("solid", fgColor=bg)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 38

    def set_w(ws, cols, widths):
        for ci, cn in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(cn, 12)

    def signal_style(sig):
        s = str(sig)
        if "STRONG" in s: return STAR3_BG, GREEN_C
        if "GOOD"   in s: return STAR2_BG, MOD_C
        if "OVERBOUGHT" in s: return ORANGE, SOFT_C
        if "ABOVE 200"  in s: return GOLD,   MOD_C
        if "BELOW"      in s: return RED_BG,  RED_C
        if "NO DATA"    in s: return LGREY,  "7F7F7F"
        return WHITE, "000000"

    col_w = {
        "Symbol": 12, "Company Name": 26, "Index Category": 14,
        "Cap Type": 9, "Sector": 20,
        "Price": 10, "SMA 50": 10, "SMA 100": 10, "SMA 200": 10,
        "% vs SMA 50": 11, "% vs SMA 100": 12, "% vs SMA 200": 12,
        "Above 50": 9, "Above 100": 9, "Above 200": 9, "Within 10%": 10,
        "Sweet Spot": 11, "SMA Signal": 18,
        "Data Points": 10, "Scored On": 12,
        "Last Close": 10, "Stop Loss %": 10, "Stop Price": 10,
        "CAR (Day 1-10)": 50,
    }

    main_cols = [c for c in [
        "Symbol", "Company Name", "Index Category", "Cap Type", "Sector",
        "Price", "SMA 50", "SMA 100", "SMA 200",
        "% vs SMA 50", "% vs SMA 100", "% vs SMA 200",
        "Above 50", "Above 100", "Above 200", "Within 10%",
        "Sweet Spot", "SMA Signal", "Scored On"
    ] if c in scores_df.columns]

    # Subsets
    sweet_df     = scores_df[scores_df["Sweet Spot"].str.contains("YES", na=False)].copy()
    below200_df  = scores_df[scores_df["Above 200"] == "❌"].copy()
    overbought_df= scores_df[
        (scores_df["Above 200"] == "✅") &
        (scores_df["Within 10%"] == "❌")
    ].copy()

    # Sort sweet spot by % vs SMA 200 (closest to 200 SMA first = freshest breakout)
    if "% vs SMA 200" in sweet_df.columns:
        sweet_df["_diff200"] = pd.to_numeric(
            sweet_df["% vs SMA 200"], errors="coerce")
        sweet_df = sweet_df.sort_values("_diff200")
        sweet_df.drop(columns=["_diff200"], inplace=True, errors="ignore")

    scored_on = (scores_df["Scored On"].iloc[0]
                 if "Scored On" in scores_df.columns else str(date.today()))

    # ── Sheet 1: All Stocks SMA ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Stocks SMA"
    add_title(ws1,
              f"🇺🇸 USA SMA FILTER  |  "
              f"Sweet Spot:{len(sweet_df)}  "
              f"Overbought:{len(overbought_df)}  "
              f"Below 200SMA:{len(below200_df)}  |  "
              f"Scored:{scored_on}",
              len(main_cols))

    for ci, cn in enumerate(main_cols, 1):
        hdr(ws1.cell(2, ci)); ws1.cell(2, ci).value = cn
    ws1.row_dimensions[2].height = 28

    # Sort: sweet spot first, then above200, then below
    def sort_key(row):
        sig = str(row.get("SMA Signal",""))
        if "STRONG" in sig: return 0
        if "GOOD"   in sig: return 1
        if "ABOVE 200" in sig: return 2
        if "ABOVE 100" in sig: return 3
        if "ABOVE 50"  in sig: return 4
        if "OVERBOUGHT" in sig: return 5
        return 6

    scores_sorted = scores_df.copy()
    scores_sorted["_sk"] = scores_sorted.apply(
        lambda r: sort_key(r.to_dict()), axis=1)
    scores_sorted = scores_sorted.sort_values("_sk").drop(columns=["_sk"])

    for ri, row in enumerate(
            scores_sorted[main_cols].itertuples(index=False), 3):
        ws1.append(list(row))
        ws1.row_dimensions[ri].height = 18
        rd     = row._asdict()
        sig    = str(rd.get("SMA Signal",""))
        row_bg = LGREY if ri%2==0 else WHITE

        for ci, cn in enumerate(main_cols, 1):
            cell = ws1.cell(ri, ci)
            val  = rd.get(cn, None)

            if cn == "SMA Signal":
                sbg, scol = signal_style(sig)
                dat(cell, bg=sbg, bold=True, color=scol)
            elif cn == "Sweet Spot":
                dat(cell,
                    bg=STAR3_BG if "YES" in str(val) else RED_BG,
                    bold="YES" in str(val),
                    color=GREEN_C if "YES" in str(val) else RED_C)
            elif cn in ("Above 50", "Above 100", "Above 200", "Within 10%"):
                dat(cell,
                    bg=GREEN_BG if str(val)=="✅" else
                       RED_BG   if str(val)=="❌" else LGREY,
                    bold=True,
                    color=GREEN_C if str(val)=="✅" else RED_C)
            elif cn in ("% vs SMA 50","% vs SMA 100","% vs SMA 200"):
                try:
                    v = float(val)
                    dat(cell,
                        bg=(STAR3_BG if 0 < v <= 5 else
                            STAR2_BG if 0 < v <= 10 else
                            ORANGE   if v > 10 else
                            RED_BG   if v < 0 else row_bg),
                        bold=True,
                        color=GREEN_C if v > 0 else RED_C)
                except Exception:
                    dat(cell, bg=row_bg)
            elif cn == "Price":
                dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
            elif cn in ("SMA 50","SMA 100","SMA 200"):
                dat(cell, bg=row_bg, color=MID_BLUE)
            elif cn == "Company Name":
                dat(cell, bg=row_bg, align=lft)
            elif cn == "Symbol":
                dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
            else:
                dat(cell, bg=row_bg)

    ws1.freeze_panes = "A3"
    set_w(ws1, main_cols, col_w)

    # ── Sheet 2: SMA Sweet Spot ───────────────────────────────────────────
    ws2 = wb.create_sheet("SMA Sweet Spot")
    add_title(ws2,
              f"SMA SWEET SPOT  |  {len(sweet_df)} stocks  |  "
              f"Above 50+100+200 SMA AND within 10% of 200 SMA  |  "
              f"Sorted: closest to 200 SMA first",
              len(main_cols), bg=GREEN_C)

    for ci, cn in enumerate(main_cols, 1):
        hdr(ws2.cell(2, ci), bg=GREEN_C); ws2.cell(2, ci).value = cn
    ws2.row_dimensions[2].height = 28

    for ri, row in enumerate(sweet_df[main_cols].itertuples(index=False), 3):
        ws2.append(list(row))
        ws2.row_dimensions[ri].height = 18
        rd     = row._asdict()
        sig    = str(rd.get("SMA Signal",""))
        row_bg = STAR3_BG if "STRONG" in sig else STAR2_BG

        for ci, cn in enumerate(main_cols, 1):
            cell = ws2.cell(ri, ci)
            val  = rd.get(cn, None)
            if cn == "SMA Signal":
                sbg, scol = signal_style(sig)
                dat(cell, bg=sbg, bold=True, color=scol)
            elif cn == "% vs SMA 200":
                try:
                    v = float(val)
                    dat(cell, bg=STAR3_BG if v<=5 else STAR2_BG,
                        bold=True, color=GREEN_C)
                except Exception:
                    dat(cell, bg=row_bg)
            elif cn in ("% vs SMA 50","% vs SMA 100"):
                dat(cell, bg=row_bg, color=GREEN_C)
            elif cn in ("Above 50","Above 100","Above 200","Within 10%","Sweet Spot"):
                dat(cell, bg=GREEN_BG, bold=True, color=GREEN_C)
            elif cn == "Price":
                dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
            elif cn in ("SMA 50","SMA 100","SMA 200"):
                dat(cell, bg=row_bg, color=MID_BLUE)
            elif cn == "Company Name":
                dat(cell, bg=row_bg, align=lft)
            elif cn == "Symbol":
                dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
            else:
                dat(cell, bg=row_bg)

    ws2.freeze_panes = "A3"
    set_w(ws2, main_cols, col_w)

    # ── Sheet 3: Below 200 SMA ────────────────────────────────────────────
    ws3 = wb.create_sheet("Below 200 SMA")
    add_title(ws3,
              f"BELOW 200 SMA — AVOID  |  {len(below200_df)} stocks  |  "
              f"Downtrend — skip CAR signals from these!",
              len(main_cols), bg=RED_C)

    for ci, cn in enumerate(main_cols, 1):
        hdr(ws3.cell(2, ci), bg=RED_C); ws3.cell(2, ci).value = cn
    ws3.row_dimensions[2].height = 28

    for ri, row in enumerate(below200_df[main_cols].itertuples(index=False), 3):
        ws3.append(list(row))
        ws3.row_dimensions[ri].height = 18
        rd = row._asdict()
        for ci, cn in enumerate(main_cols, 1):
            cell   = ws3.cell(ri, ci)
            val    = rd.get(cn, None)
            row_bg = LGREY if ri%2==0 else WHITE
            if cn == "SMA Signal":
                dat(cell, bg=RED_BG, bold=True, color=RED_C)
            elif cn in ("Above 50","Above 100","Above 200","Within 10%","Sweet Spot"):
                dat(cell,
                    bg=RED_BG if str(val) in ("❌","❌ NO") else row_bg,
                    color=RED_C)
            elif cn == "% vs SMA 200":
                try:
                    v = float(val)
                    dat(cell, bg=RED_BG, bold=True, color=RED_C)
                except Exception:
                    dat(cell, bg=row_bg)
            elif cn == "Company Name":
                dat(cell, bg=row_bg, align=lft)
            elif cn == "Symbol":
                dat(cell, bg=row_bg, bold=True, color=RED_C)
            else:
                dat(cell, bg=row_bg)

    ws3.freeze_panes = "A3"
    set_w(ws3, main_cols, col_w)

    # ── Sheet 4: Overbought >10% ──────────────────────────────────────────
    ws4 = wb.create_sheet("Overbought >10%")
    add_title(ws4,
              f"OVERBOUGHT >10% above 200 SMA  |  {len(overbought_df)} stocks  |  "
              f"Above 200 SMA but too extended — wait for pullback!",
              len(main_cols), bg=SOFT_C)

    for ci, cn in enumerate(main_cols, 1):
        hdr(ws4.cell(2, ci), bg=SOFT_C); ws4.cell(2, ci).value = cn
    ws4.row_dimensions[2].height = 28

    if not overbought_df.empty:
        if "% vs SMA 200" in overbought_df.columns:
            overbought_df["_d"] = pd.to_numeric(
                overbought_df["% vs SMA 200"], errors="coerce")
            overbought_df = overbought_df.sort_values("_d", ascending=False)
            overbought_df.drop(columns=["_d"], inplace=True, errors="ignore")

        for ri, row in enumerate(
                overbought_df[main_cols].itertuples(index=False), 3):
            ws4.append(list(row))
            ws4.row_dimensions[ri].height = 18
            rd     = row._asdict()
            row_bg = LGREY if ri%2==0 else WHITE
            for ci, cn in enumerate(main_cols, 1):
                cell = ws4.cell(ri, ci)
                val  = rd.get(cn, None)
                if cn == "SMA Signal":
                    dat(cell, bg=ORANGE, bold=True, color=SOFT_C)
                elif cn == "% vs SMA 200":
                    try:
                        v = float(val)
                        dat(cell,
                            bg=RED_BG if v>20 else ORANGE,
                            bold=True,
                            color=RED_C if v>20 else SOFT_C)
                    except Exception:
                        dat(cell, bg=row_bg)
                elif cn == "Within 10%":
                    dat(cell, bg=ORANGE, bold=True, color=SOFT_C)
                elif cn == "Company Name":
                    dat(cell, bg=row_bg, align=lft)
                elif cn == "Symbol":
                    dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
                else:
                    dat(cell, bg=row_bg)

    ws4.freeze_panes = "A3"
    set_w(ws4, main_cols, col_w)

    # ── Sheet 5: CAR + SMA Combined ──────────────────────────────────────
    ws5 = wb.create_sheet("CAR + SMA Combined ⭐")

    if signals_df is not None and not signals_df.empty:
        # Merge CAR signals with SMA scores
        sma_cols = [c for c in ["Symbol",
                                 "% vs SMA 50","% vs SMA 100","% vs SMA 200",
                                 "Above 50","Above 100","Above 200","Within 10%",
                                 "Sweet Spot","SMA Signal",
                                 "SMA 50","SMA 100","SMA 200"]
                    if c in scores_df.columns]

        combined = signals_df.merge(
            scores_df[sma_cols], on="Symbol", how="left")

        # Filter: only sweet spot
        combined_sweet = combined[
            combined["Sweet Spot"].str.contains("YES", na=False)
        ].copy()

        # Sort by % vs SMA 200 (closest = freshest signal)
        if "% vs SMA 200" in combined_sweet.columns:
            combined_sweet["_d"] = pd.to_numeric(
                combined_sweet["% vs SMA 200"], errors="coerce")
            combined_sweet = combined_sweet.sort_values("_d")
            combined_sweet.drop(columns=["_d"], inplace=True, errors="ignore")

        combined_sweet["Rank"] = range(1, len(combined_sweet)+1)

        comb_cols = [c for c in [
            "Rank", "Symbol", "Company Name", "Cap Type", "Sector",
            "Last Close", "Stop Loss %", "Stop Price",
            "% vs SMA 50","% vs SMA 100","% vs SMA 200",
            "Above 50","Above 100","Above 200","Within 10%","SMA Signal",
            "CAR (Day 1-10)"
        ] if c in combined_sweet.columns]

        add_title(ws5,
                  f"⭐ CAR + SMA COMBINED  |  "
                  f"CAR Signals:{len(signals_df)}  →  "
                  f"After SMA Filter:{len(combined_sweet)}  |  "
                  f"BEST QUALITY PICKS!  |  {date.today()}",
                  len(comb_cols), bg=TEAL)

        for ci, cn in enumerate(comb_cols, 1):
            hdr(ws5.cell(2, ci), bg=TEAL); ws5.cell(2, ci).value = cn
        ws5.row_dimensions[2].height = 28

        for ri, row in enumerate(
                combined_sweet[comb_cols].itertuples(index=False), 3):
            ws5.append(list(row))
            ws5.row_dimensions[ri].height = 18
            rd     = row._asdict()
            sig    = str(rd.get("SMA Signal",""))
            row_bg = STAR3_BG if "STRONG" in sig else STAR2_BG

            for ci, cn in enumerate(comb_cols, 1):
                cell = ws5.cell(ri, ci)
                val  = rd.get(cn, None)

                if cn == "SMA Signal":
                    sbg, scol = signal_style(sig)
                    dat(cell, bg=sbg, bold=True, color=scol)
                elif cn == "Rank":
                    try:
                        v = int(float(val))
                    except Exception:
                        v = 99
                    dat(cell, bg=STAR3_BG if v<=5 else row_bg,
                        bold=v<=5, color=GREEN_C)
                elif cn in ("% vs SMA 50","% vs SMA 100","% vs SMA 200"):
                    try:
                        v = float(val)
                        dat(cell,
                            bg=STAR3_BG if 0<v<=5 else STAR2_BG if 0<v<=10 else row_bg,
                            bold=True, color=GREEN_C)
                    except Exception:
                        dat(cell, bg=row_bg)
                elif cn in ("Above 50","Above 100","Above 200","Within 10%"):
                    dat(cell, bg=GREEN_BG, bold=True, color=GREEN_C)
                elif cn == "Last Close":
                    dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
                elif cn == "Stop Price":
                    dat(cell, bg=RED_BG, bold=True, color=RED_C)
                elif cn == "CAR (Day 1-10)":
                    dat(cell, bg="F2F9EE", align=lft)
                elif cn == "Company Name":
                    dat(cell, bg=row_bg, align=lft)
                elif cn == "Symbol":
                    dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
                else:
                    dat(cell, bg=row_bg)

        ws5.freeze_panes = "A3"
        set_w(ws5, comb_cols, col_w)

        print(f"\n  CAR Signals today  : {len(signals_df)}")
        print(f"  After SMA filter   : {len(combined_sweet)} stocks")
        print(f"  Filtered out       : {len(signals_df)-len(combined_sweet)} stocks")

    else:
        ws5.cell(2, 1).value = (
            f"No CAR signals found — "
            f"run car_signal_scanner.py first, "
            f"then re-run sma_filter.py")
        add_title(ws5, "CAR + SMA Combined — Run scanner first!", 3, bg=TEAL)
        print("\n  No CAR signals file — Sheet 5 empty")

    # ── Sheet 6: How to Use ───────────────────────────────────────────────
    ws6 = wb.create_sheet("How to Use")
    ws6.column_dimensions["A"].width = 24
    ws6.column_dimensions["B"].width = 55
    add_title(ws6, "HOW TO USE — SMA Filter Guide", 2, bg=DARK_BLUE)

    guide = [
        ("WHAT IS SMA?", ""),
        ("SMA 50",   "Average of last 50 days closing price"),
        ("SMA 100",  "Average of last 100 days closing price"),
        ("SMA 200",  "Average of last 200 days — THE most important!"),
        ("", ""),
        ("THE CONDITIONS", ""),
        ("Condition 1", "Price > 50 SMA  (short term uptrend)"),
        ("Condition 2", "Price > 100 SMA (medium term uptrend)"),
        ("Condition 3", "Price > 200 SMA (long term uptrend)"),
        ("Condition 4", "Price NOT more than 10% above 200 SMA"),
        ("Sweet Spot",  "ALL 4 conditions met = BEST buying zone!"),
        ("", ""),
        ("WHY 10% LIMIT?", ""),
        ("Too far above",  "Stock is extended / overbought — risky to buy"),
        ("Within 10%",     "Fresh breakout OR returning after pullback"),
        ("Best zone",      "0-5% above 200 SMA = STRONG signal"),
        ("Good zone",      "5-10% above 200 SMA = GOOD signal"),
        ("", ""),
        ("DAILY ROUTINE", ""),
        ("Step 1", "python us_car_signal_scanner.py  (after 4 PM EST)"),
        ("Step 2", "python us_sma_filter.py          (immediately after)"),
        ("Step 3", "Open us_us_sma_filter_report.xlsx"),
        ("Step 4", "Go to 'CAR + SMA Combined' sheet"),
        ("Step 5", "These are your BEST QUALITY picks for today!"),
        ("", ""),
        ("SHEETS GUIDE", ""),
        ("All Stocks SMA",    "All 500+ stocks with SMA values"),
        ("SMA Sweet Spot",    "Only stocks in ideal buying zone"),
        ("Below 200 SMA",     "Stocks in downtrend — AVOID"),
        ("Overbought >10%",   "Too extended — wait for pullback"),
        ("CAR + SMA Combined","CAR signal + SMA filter = BEST picks ⭐"),
        ("", ""),
        ("FILES", ""),
        ("us_sma_scores.csv",       "Saved today — reuse without re-downloading"),
        ("us_sma_filter_report.xlsx","Main report — open this!"),
    ]

    for ri, (label, value) in enumerate(guide, 2):
        ws6.row_dimensions[ri].height = 22
        is_section = (value == "")
        for ci, val in enumerate([label, value], 1):
            cell        = ws6.cell(ri, ci)
            cell.value  = val if not is_section else (label if ci==1 else "")
            cell.border = border
            if is_section:
                cell.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
                cell.fill      = PatternFill("solid", fgColor=MID_BLUE)
                cell.alignment = lft
            else:
                row_bg         = LGREY if ri%2==0 else WHITE
                cell.fill      = PatternFill("solid", fgColor=row_bg)
                cell.alignment = lft
                cell.font      = Font(bold=(ci==1),
                                      color=DARK_BLUE if ci==1 else "000000",
                                      name="Arial", size=10)

    wb.save(OUTPUT_FILE)

    # Print summary
    print(f"\n{'='*60}")
    print(f"  Saved  -> '{OUTPUT_FILE}'")
    print(f"  Scores -> '{SCORES_CSV}'")
    print(f"\n  SMA FILTER SUMMARY:")
    print(f"  Sweet Spot (all 4 conditions) : {len(sweet_df):>4} stocks")
    print(f"  Overbought (>10% above 200)   : {len(overbought_df):>4} stocks")
    print(f"  Below 200 SMA (downtrend)     : {len(below200_df):>4} stocks")
    print(f"{'='*60}")


def main():
    print("SMA Filter — USA")
    print("=" * 50)
    print(f"Date: {date.today()}\n")

    if not os.path.exists(MASTER_FILE):
        print(f"ERROR: '{MASTER_FILE}' not found!")
        print("Run build_master_ref.py first.")
        return

    master_df = pd.read_csv(MASTER_FILE)
    print(f"Loaded {len(master_df)} stocks from '{MASTER_FILE}'")

    # Check if today's SMA scores already exist
    if os.path.exists(SCORES_CSV):
        existing  = pd.read_csv(SCORES_CSV)
        scored_on = (existing["Scored On"].iloc[0]
                     if "Scored On" in existing.columns else "")
        if scored_on == str(date.today()):
            print(f"Today's SMA scores already in '{SCORES_CSV}' — skipping fetch!")
            scores_df = existing
        else:
            print(f"SMA scores from {scored_on} — fetching fresh data...")
            scores_df = build_sma_scores(master_df)
            scores_df.to_csv(SCORES_CSV, index=False)
    else:
        scores_df = build_sma_scores(master_df)
        scores_df.to_csv(SCORES_CSV, index=False)

    print(f"\nSMA scores saved -> '{SCORES_CSV}'")

    # Load CAR signals if available
    signals_df = None
    if os.path.exists(SIGNALS_FILE):
        try:
            signals_df = pd.read_excel(SIGNALS_FILE)
            if "Symbol" not in signals_df.columns:
                signals_df = pd.read_excel(SIGNALS_FILE, header=1)
            signals_df["Symbol"] = (signals_df["Symbol"]
                                    .astype(str).str.strip().str.upper())
            print(f"CAR signals loaded: {len(signals_df)} signals")
        except Exception as e:
            print(f"CAR signals skipped: {e}")

    print("\nWriting Excel report...")
    write_excel(scores_df, signals_df)


if __name__ == "__main__":
    main()
