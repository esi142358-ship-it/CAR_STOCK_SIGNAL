"""
NSE Stock Buy Signal Scanner
Logic mirrors the Google Sheets script:
  1. Find the 52-week high date for each stock
  2. Fetch daily close prices from that high date to today
  3. Compute the cumulative average over those closes
  4. If the last 10 cumulative averages are strictly increasing -> BUY signal

Requirements:
    pip install "yfinance>=0.2.40" pandas openpyxl requests

Usage:
    - Put your stock symbols (NSE tickers, one per line) in 'stocks.txt'
    - Run: python stock_signal_scanner.py
    - Results saved to 'buy_signals_report.xlsx'
"""

import yfinance as yf
import pandas as pd
import requests
from datetime import date, timedelta
import time
import os
import tempfile

# ── Fix yfinance SQLite cache issue on bulk runs ──────────────────────────────
# Route the tz cache to a fresh temp directory so it never collides/corrupts
_cache_dir = tempfile.mkdtemp(prefix="yf_cache_")
try:
    yf.set_tz_cache_location(_cache_dir)
except Exception:
    pass  # safe to ignore on older yfinance builds

# ── CONFIG ────────────────────────────────────────────────────────────────────
STOCKS_FILE          = "stocks.txt"
OUTPUT_FILE          = "buy_signals_report.xlsx"
DELAY_BETWEEN_STOCKS = 1.0   # seconds; raise to 1.5 if Yahoo throttles you
MAX_RETRIES          = 3
RETRY_DELAY          = 6.0   # seconds to wait before each retry
# ─────────────────────────────────────────────────────────────────────────────


def load_stocks():
    if os.path.exists(STOCKS_FILE):
        with open(STOCKS_FILE) as f:
            stocks = [line.strip().upper() for line in f if line.strip()]
        print(f"Loaded {len(stocks)} stocks from {STOCKS_FILE}")
        return stocks
    demo = ["RELIANCE", "TCS", "INFY", "HDFCBANK", "ICICIBANK",
            "WIPRO", "HINDUNILVR", "ITC", "SBIN", "BAJFINANCE"]
    print(f"'{STOCKS_FILE}' not found. Using {len(demo)} demo stocks.")
    return demo


def nse_ticker(symbol: str) -> str:
    symbol = symbol.upper()
    if symbol.endswith(".NS") or symbol.endswith(".BO"):
        return symbol
    return symbol + ".NS"


def get_52w_high_date(hist: pd.DataFrame):
    if hist.empty:
        return None
    idx = hist["Close"].idxmax()
    return idx.date() if idx is not None else None


def cumulative_averages(series: pd.Series) -> list:
    values = series.dropna().tolist()
    avgs, total = [], 0.0
    for i, v in enumerate(values, 1):
        total += v
        avgs.append(total / i)
    return avgs


def is_strictly_increasing(seq: list) -> bool:
    return all(seq[i] < seq[i + 1] for i in range(len(seq) - 1))


def fetch_with_retry(ticker: str, start: date, today: date):
    """Download with retries. Uses a fresh requests Session to avoid stale connections."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            # Create a fresh Ticker object each attempt to avoid cached broken state
            t = yf.Ticker(ticker)
            data = t.history(
                start=start,
                end=today + timedelta(days=1),
                auto_adjust=True,
                actions=False,
            )
            return data

        except Exception as e:
            err = str(e)
            is_retryable = any(kw in err for kw in [
                "NoneType", "unable to open database", "OperationalError",
                "stat: path", "Connection", "RemoteDisconnected", "Timeout"
            ])
            if is_retryable and attempt < MAX_RETRIES:
                print(f"  [RETRY {attempt}/{MAX_RETRIES}] {ticker} – {err[:60]}... waiting {RETRY_DELAY}s")
                time.sleep(RETRY_DELAY)
            else:
                print(f"  [ERROR] {ticker} – {err[:80]}")
                return None
    return None


def scan_stock(symbol: str):
    ticker = nse_ticker(symbol)
    today  = date.today()
    start  = today - timedelta(days=365)

    data = fetch_with_retry(ticker, start, today)

    if data is None or data.empty or len(data) < 11:
        rows = 0 if data is None else len(data)
        print(f"  [SKIP]  {ticker} – insufficient data ({rows} rows)")
        return None

    # Flatten MultiIndex columns if present (yfinance >= 0.2.x sometimes adds them)
    if isinstance(data.columns, pd.MultiIndex):
        data.columns = data.columns.get_level_values(0)

    high_date = get_52w_high_date(data)
    if high_date is None:
        return None

    closes_from_high = data.loc[data.index.date >= high_date, "Close"]
    if len(closes_from_high) < 10:
        print(f"  [SKIP]  {ticker} – only {len(closes_from_high)} closes since 52w high")
        return None

    cum_avgs = cumulative_averages(closes_from_high)
    if len(cum_avgs) < 10:
        return None

    last10 = cum_avgs[-10:]
    if is_strictly_increasing(last10):
        # Get the actual dates for the last 10 trading days
        last10_dates = closes_from_high.index[-10:]
        result = {
            "Ticker":        ticker,
            "Symbol":        symbol,
            "Signal Date":   str(today),
            "52W High Date": str(high_date),
            "Signal":        "BUY / AVERAGE OUT",
        }
        # CAR Day 1..10 columns with actual date label
        for i, (d, val) in enumerate(zip(last10_dates, last10), 1):
            col = f"CAR Day {i} ({d.strftime('%d-%b')})"
            result[col] = round(val, 2)
        return result
    return None


def main():
    stocks  = load_stocks()
    results = []
    total   = len(stocks)

    print(f"\nScanning {total} stocks...\n{'─'*50}")

    for i, symbol in enumerate(stocks, 1):
        print(f"[{i:>3}/{total}] {symbol:<20}", end="", flush=True)
        result = scan_stock(symbol)
        if result:
            results.append(result)
            print("BUY SIGNAL")
        else:
            print("-")
        time.sleep(DELAY_BETWEEN_STOCKS)

    print(f"\n{'─'*50}")
    print(f"Done. {len(results)} BUY signal(s) found out of {total} stocks.\n")

    if results:
        df = pd.DataFrame(results)

        # ── Console print ──────────────────────────────────────────────────
        summary_cols = ["Ticker", "Signal Date", "52W High Date", "Signal"]
        car_cols     = [c for c in df.columns if c.startswith("CAR Day")]
        print("\nBUY SIGNALS SUMMARY")
        print(df[summary_cols].to_string(index=False))
        print("\nLAST 10-DAY CAR VALUES")
        print(df[["Ticker"] + car_cols].to_string(index=False))

        # ── Excel output with formatting ───────────────────────────────────
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Buy Signals Report"

        header_fill   = PatternFill("solid", fgColor="1F4E79")
        car_fill      = PatternFill("solid", fgColor="E2EFDA")  # light green for CAR cols
        header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        data_font     = Font(name="Arial", size=10)
        signal_font   = Font(bold=True, color="006100", name="Arial", size=10)
        thin          = Side(style="thin", color="BFBFBF")
        border        = Border(left=thin, right=thin, top=thin, bottom=thin)
        center        = Alignment(horizontal="center", vertical="center")

        columns = list(df.columns)
        ws.append(columns)

        # Style header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill   = car_fill if col_name.startswith("CAR") else header_fill
            cell.font   = Font(bold=True, color="375623" if col_name.startswith("CAR") else "FFFFFF",
                               name="Arial", size=10)
            cell.alignment = center
            cell.border = border

        # Write data rows
        for row_data in df.itertuples(index=False):
            ws.append(list(row_data))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_name = columns[cell.column - 1]
                cell.font      = signal_font if col_name == "Signal" else data_font
                cell.alignment = center
                cell.border    = border
                if col_name.startswith("CAR"):
                    cell.fill = PatternFill("solid", fgColor="F2F9EE")

        # Auto-fit column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_len = max(len(str(col_name)),
                         max((len(str(v)) for v in df.iloc[:, col_idx-1]), default=0))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        wb.save(OUTPUT_FILE)
        print(f"\nResults saved to '{OUTPUT_FILE}' ({len(results)} stock(s))")
    else:
        print("No BUY signals found.")


if __name__ == "__main__":
    main()
