"""
SMA Filter — India CAR Signals
================================
Runs ONLY on today's CAR signals (fast — 2-3 min!)

Condition (exact):
  Price > 50 SMA
  Price > 100 SMA
  Price > 200 SMA
  Price between 200 SMA and 200 SMA x 1.10

Reads:
  buy_signals_report.xlsx  <- today's CAR signals (required)
  nse_master_reference.csv <- sector/cap info (optional)

Creates: sma_filter_report.xlsx

Sheets:
  1. All CAR Signals + SMA   -- all signals with SMA values
  2. SMA Sweet Spot          -- stocks meeting ALL 4 conditions
  3. Failed SMA Filter       -- signals that did not pass
  4. CAR + SMA Combined      -- BEST quality picks (use this!)
  5. How to Use

Usage:
  python sma_filter.py

Run AFTER car_signal_scanner.py !
"""

import yfinance as yf
import pandas as pd
import shutil
import os, time, warnings, tempfile
from datetime import date, timedelta
warnings.filterwarnings("ignore")

_cache_dir = tempfile.mkdtemp(prefix="yf_sma_")
try:
    yf.set_tz_cache_location(_cache_dir)
except Exception:
    pass

SIGNALS_FILE  = "buy_signals_report.xlsx"
MASTER_FILE   = "nse_master_reference.csv"
OUTPUT_FILE   = "sma_filter_report.xlsx"
today = date.today().strftime("%Y-%m-%d")
DATED_FILE = f"sma_filter_report_{today}.xlsx"
DELAY         = 0.5
MAX_RETRIES   = 3
SMA_MAX_ABOVE = 1.10   # 10% above 200 SMA = upper limit


def nse_ticker(symbol):
    return symbol + ".NS"


def fetch_sma(symbol):
    ticker = nse_ticker(symbol)
    end    = date.today()
    start  = end - timedelta(days=420)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            hist = yf.Ticker(ticker).history(
                start=str(start), end=str(end),
                auto_adjust=True, actions=False)

            if hist is None or hist.empty:
                return None

            if isinstance(hist.columns, pd.MultiIndex):
                hist.columns = hist.columns.get_level_values(0)

            closes = hist["Close"].dropna()
            if len(closes) < 50:
                return None

            curr   = round(float(closes.iloc[-1]), 2)
            sma50  = round(float(closes.tail(50).mean()),  2) if len(closes) >= 50  else None
            sma100 = round(float(closes.tail(100).mean()), 2) if len(closes) >= 100 else None
            sma200 = round(float(closes.tail(200).mean()), 2) if len(closes) >= 200 else None

            def pct(price, sma):
                if sma and sma > 0:
                    return round((price - sma) / sma * 100, 2)
                return None

            diff50  = pct(curr, sma50)
            diff100 = pct(curr, sma100)
            diff200 = pct(curr, sma200)

            above50  = sma50  is not None and curr > sma50
            above100 = sma100 is not None and curr > sma100
            above200 = sma200 is not None and curr > sma200
            within10 = sma200 is not None and curr <= sma200 * SMA_MAX_ABOVE

            sweet = above50 and above100 and above200 and within10

            if sweet:
                signal = "STRONG" if (diff200 is not None and diff200 <= 5) else "GOOD"
            elif above200 and not within10:
                signal = "OVERBOUGHT"
            elif above100:
                signal = "BELOW 200 SMA"
            elif above50:
                signal = "BELOW 100 SMA"
            else:
                signal = "BELOW ALL SMAs"

            signal_label = {
                "STRONG":       "STRONG 0-5% above 200SMA",
                "GOOD":         "GOOD 5-10% above 200SMA",
                "OVERBOUGHT":   "OVERBOUGHT >10% above 200SMA",
                "BELOW 200 SMA":"BELOW 200 SMA",
                "BELOW 100 SMA":"BELOW 100 SMA",
                "BELOW ALL SMAs":"BELOW ALL SMAs",
            }.get(signal, signal)

            return {
                "Current Price":  curr,
                "SMA 50":         sma50,
                "SMA 100":        sma100,
                "SMA 200":        sma200,
                "% vs SMA 50":    diff50,
                "% vs SMA 100":   diff100,
                "% vs SMA 200":   diff200,
                "Above 50 SMA":   "YES" if above50  else "NO",
                "Above 100 SMA":  "YES" if above100 else "NO",
                "Above 200 SMA":  "YES" if above200 else "NO",
                "Within 10%":     "YES" if (within10 and above200) else "NO",
                "Sweet Spot":     "YES" if sweet else "NO",
                "SMA Signal":     signal_label,
                "_signal_key":    signal,
            }

        except Exception:
            if attempt < MAX_RETRIES:
                time.sleep(1)
            else:
                return None
    return None


def run_sma_on_signals(signals_df, master_df):
    total   = len(signals_df)
    records = []

    print(f"\nFetching SMA for {total} CAR signals (~{total} x 0.5s)...\n")

    for i, (_, row) in enumerate(signals_df.iterrows(), 1):
        rd     = row.to_dict()
        symbol = str(rd.get("Symbol", "")).strip().upper()

        print(f"  [{i:>2}/{total}] {symbol:<12}", end="", flush=True)

        sma = fetch_sma(symbol)

        if sma:
            diff   = sma["% vs SMA 200"]
            ds     = f"{diff:+.1f}%" if diff is not None else "?"
            print(f" Price:Rs{sma['Current Price']:>7}  "
                  f"SMA200:Rs{sma['SMA 200']:>7}  "
                  f"Diff:{ds:>7}  [{sma['SMA Signal']}]")
        else:
            print("  no data")
            sma = {
                "Current Price": None, "SMA 50": None,
                "SMA 100": None, "SMA 200": None,
                "% vs SMA 50": None, "% vs SMA 100": None,
                "% vs SMA 200": None,
                "Above 50 SMA": "?", "Above 100 SMA": "?",
                "Above 200 SMA": "?", "Within 10%": "?",
                "Sweet Spot": "NO", "SMA Signal": "NO DATA",
                "_signal_key": "NO DATA",
            }

        record = {}
        record.update(rd)
        record.update(sma)
        records.append(record)
        time.sleep(DELAY)

    df = pd.DataFrame(records)

    if master_df is not None:
        m_cols = [c for c in ["Symbol", "Company Name",
                               "Cap Type", "Sector", "Index Category"]
                  if c in master_df.columns]
        df = df.merge(master_df[m_cols], on="Symbol",
                      how="left", suffixes=("", "_m"))
        for col in ["Company Name", "Cap Type", "Sector", "Index Category"]:
            if col + "_m" in df.columns:
                df[col] = df[col].fillna(df[col + "_m"])
                df.drop(columns=[col + "_m"], inplace=True, errors="ignore")

    return df


def write_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill,
                                  Alignment, Border, Side)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Palette
    DARK_BLUE = "1F4E79"; MID_BLUE = "2E75B6"; TEAL    = "1F7391"
    WHITE     = "FFFFFF"; LGREY   = "F5F5F5"
    GREEN_BG  = "E2EFDA"; RED_BG  = "FFE2E2"
    GOLD      = "FFF2CC"; ORANGE  = "FCE4D6"
    GREEN_C   = "375623"; RED_C   = "C00000"
    MOD_C     = "7F6000"; SOFT_C  = "833C00"
    STAR3_BG  = "C6EFCE"; STAR2_BG= "FFEB9C"

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft    = Alignment(horizontal="left",   vertical="center")

    def hdr(cell, bg=DARK_BLUE, fg=WHITE):
        cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = ctr; cell.border = border

    def dat(cell, bg=None, bold=False, color="000000", align=None):
        cell.font      = Font(bold=bold, color=color, name="Arial", size=10)
        cell.alignment = align if align else ctr
        cell.border    = border
        if bg: cell.fill = PatternFill("solid", fgColor=bg)

    def title(ws, text, ncols, bg=DARK_BLUE, fg=WHITE):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value     = text
        c.font      = Font(bold=True, color=fg, name="Arial", size=13)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 38

    def widths(ws, cols, cw):
        for ci, cn in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = cw.get(cn, 12)

    def sig_style(sig):
        if "STRONG"     in sig: return STAR3_BG, GREEN_C
        if "GOOD"       in sig: return STAR2_BG, MOD_C
        if "OVERBOUGHT" in sig: return ORANGE,   SOFT_C
        if "BELOW"      in sig: return RED_BG,   RED_C
        return LGREY, "7F7F7F"

    CW = {
        "Symbol":11, "Company Name":26, "Index Category":13,
        "Cap Type":9, "Sector":20,
        "Signal Date":12, "52W High Date":13,
        "Current Price":13, "SMA 50":10, "SMA 100":10, "SMA 200":11,
        "% vs SMA 50":11, "% vs SMA 100":12, "% vs SMA 200":12,
        "Above 50 SMA":10, "Above 100 SMA":11, "Above 200 SMA":11,
        "Within 10%":10, "Sweet Spot":11, "SMA Signal":28,
        "Stop Loss %":10, "Stop Price":11, "Rank":6,
    }

    # CAR day columns
    car_cols = [c for c in df.columns
                if "CAR" in str(c) or "Signal" in str(c)
                or "52W" in str(c)]

    front_cols = [c for c in
                  ["Symbol","Company Name","Cap Type","Sector",
                   "Index Category","Current Price",
                   "SMA 50","SMA 100","SMA 200",
                   "% vs SMA 50","% vs SMA 100","% vs SMA 200",
                   "Above 50 SMA","Above 100 SMA","Above 200 SMA",
                   "Within 10%","Sweet Spot","SMA Signal"]
                  if c in df.columns]

    all_cols = front_cols + [c for c in car_cols if c not in front_cols]

    sweet = df[df["Sweet Spot"] == "YES"].copy()
    fail  = df[df["Sweet Spot"] != "YES"].copy()

    if "% vs SMA 200" in sweet.columns:
        sweet["_d"] = pd.to_numeric(sweet["% vs SMA 200"], errors="coerce")
        sweet = sweet.sort_values("_d")
        sweet.drop(columns=["_d"], inplace=True, errors="ignore")

    overbought = fail[fail["SMA Signal"].str.contains("OVERBOUGHT", na=False)]
    below_sma  = fail[~fail["SMA Signal"].str.contains("OVERBOUGHT", na=False)]

    def style_row(ws, ri, rd, cols, force_bg=None):
        sig    = str(rd.get("SMA Signal",""))
        row_bg = force_bg if force_bg else (LGREY if ri%2==0 else WHITE)
        for ci, cn in enumerate(cols, 1):
            cell = ws.cell(ri, ci)
            val  = rd.get(cn, None)
            if cn == "SMA Signal":
                sb, sc = sig_style(sig)
                dat(cell, bg=sb, bold=True, color=sc)
            elif cn == "Sweet Spot":
                is_y = str(val) == "YES"
                dat(cell, bg=GREEN_BG if is_y else RED_BG,
                    bold=True, color=GREEN_C if is_y else RED_C,
                    align=ctr)
                cell.value = "YES" if is_y else "NO"
            elif cn in ("Above 50 SMA","Above 100 SMA",
                        "Above 200 SMA","Within 10%"):
                is_y = str(val) == "YES"
                dat(cell, bg=GREEN_BG if is_y else RED_BG,
                    bold=True, color=GREEN_C if is_y else RED_C)
                cell.value = "YES" if is_y else "NO"
            elif cn == "% vs SMA 200":
                try:
                    v = float(val)
                    bg = (STAR3_BG if 0 < v <= 5 else
                          STAR2_BG if 0 < v <= 10 else
                          ORANGE   if v > 10 else RED_BG)
                    dat(cell, bg=bg, bold=True,
                        color=GREEN_C if v > 0 else RED_C)
                except Exception:
                    dat(cell, bg=row_bg)
            elif cn in ("% vs SMA 50","% vs SMA 100"):
                try:
                    v = float(val)
                    dat(cell, bg=GREEN_BG if v>0 else RED_BG,
                        color=GREEN_C if v>0 else RED_C)
                except Exception:
                    dat(cell, bg=row_bg)
            elif cn == "Current Price":
                dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
            elif cn in ("SMA 50","SMA 100","SMA 200"):
                dat(cell, bg=row_bg, color=MID_BLUE)
            elif cn == "Stop Price":
                dat(cell, bg=RED_BG, bold=True, color=RED_C)
            elif cn == "Rank":
                try:
                    v = int(float(val))
                    dat(cell, bg=STAR3_BG if v<=3 else STAR2_BG,
                        bold=True, color=GREEN_C)
                except Exception:
                    dat(cell, bg=row_bg)
            elif cn == "Company Name":
                dat(cell, bg=row_bg, align=lft)
            elif cn == "Symbol":
                dat(cell, bg=row_bg, bold=True, color=DARK_BLUE)
            elif any(k in str(cn) for k in ["CAR","Signal","52W"]):
                dat(cell, bg="F2F9EE", align=lft)
            else:
                dat(cell, bg=row_bg)

    def write_sheet(ws, df_s, cols, hdr_bg=DARK_BLUE):
        for ci, cn in enumerate(cols, 1):
            hdr(ws.cell(2, ci), bg=hdr_bg); ws.cell(2, ci).value = cn
        ws.row_dimensions[2].height = 28
        for ri, (_, row) in enumerate(df_s[cols].iterrows(), 3):
            ws.append([row.get(c, None) for c in cols])
            ws.row_dimensions[ri].height = 18
            style_row(ws, ri, row.to_dict(), cols)
        ws.freeze_panes = "A3"
        widths(ws, cols, CW)

    # ── Sheet 1: All CAR Signals + SMA ───────────────────────────────────
    ws1 = wb.active; ws1.title = "All CAR Signals + SMA"
    title(ws1,
          f"ALL CAR SIGNALS + SMA  |  Total:{len(df)}  "
          f"Sweet Spot:{len(sweet)}  Failed:{len(fail)}  |  {date.today()}",
          len(all_cols))
    write_sheet(ws1, df, all_cols)

    # ── Sheet 2: SMA Sweet Spot ───────────────────────────────────────────
    ws2 = wb.create_sheet("SMA Sweet Spot")
    title(ws2,
          f"SMA SWEET SPOT  |  {len(sweet)} stocks pass ALL 4 conditions  |  "
          f"Sorted: closest to 200 SMA first",
          len(all_cols), bg=GREEN_C)
    write_sheet(ws2, sweet, all_cols, hdr_bg=GREEN_C)

    # ── Sheet 3: Failed SMA Filter ────────────────────────────────────────
    ws3 = wb.create_sheet("Failed SMA Filter")
    title(ws3,
          f"FAILED SMA FILTER  |  {len(fail)} stocks  |  "
          f"Overbought:{len(overbought)}  Below SMA:{len(below_sma)}",
          len(all_cols), bg=RED_C)
    write_sheet(ws3, fail, all_cols, hdr_bg=RED_C)

    # ── Sheet 4: CAR + SMA Combined ──────────────────────────────────────
    ws4 = wb.create_sheet("CAR + SMA Combined")
    combined = sweet.copy()
    combined.insert(0, "Rank", range(1, len(combined)+1))
    comb_cols = ["Rank"] + all_cols

    title(ws4,
          f"CAR + SMA COMBINED — BEST PICKS  |  "
          f"CAR Signals:{len(df)}  After SMA Filter:{len(sweet)}  |  "
          f"USE THIS SHEET for buying!  |  {date.today()}",
          len(comb_cols), bg=TEAL)
    write_sheet(ws4, combined, comb_cols, hdr_bg=TEAL)

    # ── Sheet 5: How to Use ───────────────────────────────────────────────
    ws5 = wb.create_sheet("How to Use")
    ws5.column_dimensions["A"].width = 24
    ws5.column_dimensions["B"].width = 58
    title(ws5, "HOW TO USE  SMA Filter Guide", 2, bg=DARK_BLUE)

    guide = [
        ("WHAT IS SMA?",""),
        ("SMA 50",  "Average of last 50 days closing price (short term)"),
        ("SMA 100", "Average of last 100 days closing price (medium term)"),
        ("SMA 200", "Average of last 200 days closing price -- MOST IMPORTANT"),
        ("",""),
        ("THE 4 CONDITIONS",""),
        ("Condition 1", "Price > 50 SMA  (short term uptrend)"),
        ("Condition 2", "Price > 100 SMA (medium term uptrend)"),
        ("Condition 3", "Price > 200 SMA (long term uptrend)"),
        ("Condition 4", "Price <= 200 SMA x 1.10 (not more than 10% above 200 SMA)"),
        ("Sweet Spot",  "ALL 4 conditions pass = best buying zone!"),
        ("",""),
        ("WHY 10% LIMIT?",""),
        ("Too far above", ">10% above 200 SMA = stock is extended/overbought"),
        ("Sweet zone",    "0-5% above 200 SMA  = STRONG -- freshest signal"),
        ("Good zone",     "5-10% above 200 SMA = GOOD   -- still ok to buy"),
        ("Overbought",    ">10% above 200 SMA  = wait for pullback first!"),
        ("",""),
        ("DAILY ROUTINE",""),
        ("Step 1", "4 PM IST: python car_signal_scanner.py"),
        ("Step 2", "python sma_filter.py  (runs only on today's signals)"),
        ("Step 3", "Open sma_filter_report.xlsx"),
        ("Step 4", "Go to 'CAR + SMA Combined' sheet"),
        ("Step 5", "STRONG stocks at top = buy first!"),
        ("",""),
        ("SHEETS GUIDE",""),
        ("All CAR Signals + SMA",  "All signals with SMA values"),
        ("SMA Sweet Spot",         "Only stocks passing ALL 4 conditions"),
        ("Failed SMA Filter",      "Overbought or below SMA -- skip today"),
        ("CAR + SMA Combined",     "BEST picks -- use this for buying!"),
        ("",""),
        ("% vs SMA 200 MEANING",""),
        ("0% to +5%",    "STRONG  -- just crossed 200 SMA -- best entry!"),
        ("5% to +10%",   "GOOD    -- good entry, some upside already done"),
        ("Above +10%",   "OVERBOUGHT -- wait for pullback to 200 SMA zone"),
        ("Negative",     "BELOW 200 SMA -- downtrend -- avoid!"),
    ]

    for ri, (lbl, val) in enumerate(guide, 2):
        ws5.row_dimensions[ri].height = 22
        is_sec = (val == "")
        for ci, v in enumerate([lbl, val], 1):
            cell        = ws5.cell(ri, ci)
            cell.value  = v if not is_sec else (lbl if ci==1 else "")
            cell.border = border
            if is_sec:
                cell.font      = Font(bold=True, color=WHITE,
                                      name="Arial", size=10)
                cell.fill      = PatternFill("solid", fgColor=MID_BLUE)
                cell.alignment = lft
            else:
                bg = LGREY if ri%2==0 else WHITE
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = lft
                cell.font      = Font(bold=(ci==1),
                                      color=DARK_BLUE if ci==1 else "000000",
                                      name="Arial", size=10)

    wb.save(OUTPUT_FILE)
    shutil.copy(OUTPUT_FILE, DATED_FILE)


def print_summary(df):
    sweet = df[df["Sweet Spot"]=="YES"]
    fail  = df[df["Sweet Spot"]!="YES"]
    ovbt  = fail[fail["SMA Signal"].str.contains("OVERBOUGHT", na=False)]
    below = fail[~fail["SMA Signal"].str.contains("OVERBOUGHT", na=False)]

    print(f"\n{'='*65}")
    print(f"  SMA FILTER RESULTS -- {date.today()}")
    print(f"{'='*65}")
    print(f"  Total CAR signals today : {len(df)}")
    print(f"  {'─'*42}")
    print(f"  SWEET SPOT (BUY)        : {len(sweet)} stocks")
    print(f"  OVERBOUGHT (wait)       : {len(ovbt)} stocks")
    print(f"  BELOW SMA  (skip)       : {len(below)} stocks")

    if not sweet.empty:
        print(f"\n  TODAY'S BEST CAR + SMA PICKS:")
        print(f"  {'#':<4} {'Symbol':<11} {'Price':>9}  "
              f"{'SMA200':>9}  {'%Diff':>7}  Signal")
        print(f"  {'─'*62}")
        for rank, (_, row) in enumerate(sweet.iterrows(), 1):
            sym   = str(row.get("Symbol",""))[:10]
            price = row.get("Current Price","?")
            s200  = row.get("SMA 200","?")
            diff  = row.get("% vs SMA 200","?")
            sig   = str(row.get("SMA Signal",""))
            try:    ds = f"{float(diff):+.1f}%"
            except: ds = "?"
            print(f"  {rank:<4} {sym:<11} Rs{price!s:>8}  "
                  f"Rs{s200!s:>8}  {ds:>7}  {sig}")

    print(f"\n  Report saved: '{OUTPUT_FILE}'")
    print(f"  Open 'CAR + SMA Combined' sheet for buying!")
    print(f"{'='*65}")


def main():
    print("SMA Filter -- India NSE (Today's CAR Signals Only)")
    print("=" * 52)
    print(f"Date: {date.today()}\n")

    if not os.path.exists(SIGNALS_FILE):
        print(f"ERROR: '{SIGNALS_FILE}' not found!")
        print("Run car_signal_scanner.py first!")
        return

    try:
        signals_df = pd.read_excel(SIGNALS_FILE)
        if "Symbol" not in signals_df.columns:
            signals_df = pd.read_excel(SIGNALS_FILE, header=1)
        signals_df = signals_df.dropna(subset=["Symbol"])
        signals_df["Symbol"] = (signals_df["Symbol"]
                                .astype(str).str.strip().str.upper())
        print(f"Loaded {len(signals_df)} CAR signals")
    except Exception as e:
        print(f"ERROR: {e}"); return

    if len(signals_df) == 0:
        print("No CAR signals today -- nothing to filter!")
        return

    master_df = None
    if os.path.exists(MASTER_FILE):
        master_df = pd.read_csv(MASTER_FILE)
        master_df["Symbol"] = (master_df["Symbol"]
                               .astype(str).str.strip().str.upper())
        print(f"Loaded {len(master_df)} stocks from master reference")

    result_df = run_sma_on_signals(signals_df, master_df)
    print("\nWriting Excel report...")
    write_excel(result_df)
    print_summary(result_df)


if __name__ == "__main__":
    main()
